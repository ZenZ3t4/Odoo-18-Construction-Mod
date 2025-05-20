from funzioni_reports import *
import openpyxl
from fpdf import FPDF
from datetime import datetime
import os

# #####################################################
# -----------------------------------------
# FUNZIONI UTILIZZATE NEL CODICE
# -----------------------------------------
def stampa_dati_azienda(pdf, azienda, settore_ateco, ente_formativo, ente_certificatore):
    pdf.ln(5)
    # BLOCCO DATI AZIENDA
    if len(settore_ateco) > 80:
        settore_ateco = settore_ateco[:80] + '...'
    
    # Dati dinamici azienda
    pdf.set_font("Helvetica", "I", size=10)
    #azienda_row_1 = f"{azienda} - Settore ateco: {settore_ateco}"
    azienda_row_1 = f"{azienda}"
    pdf.cell(200, 5, txt=azienda_row_1, align = 'C', ln=True)
    #--------------------------------------------------------------------------------
    
def blocco_dati_utente(pdf, data_nascita, luogo_nascita, cognome, nome, codice_fiscale):
    # BLOCCO DATI UTENTE
    # header in grassetto
    pdf.set_font("Helvetica", "BI", size=14)
    pdf.set_margins(10, 10, 10)
    testo_rilasciato_a = "il presente attestato viene conferito a"
    pdf.cell(200, 10, txt=testo_rilasciato_a, align = 'C', ln=True)
    pdf.ln(10)
    # blocco dati dinamici
    # dati anagrafici del candidato
    nome_cognome = nome + ' ' + cognome
    dati_utente = f"nato/a a {luogo_nascita} il {data_nascita} - Codice Fiscale: {codice_fiscale}"
    
    # Stampa nome e cognome
    pdf.set_font("Helvetica", 'B', size=32)
    pdf.set_text_color(235, 9, 16)
    pdf.set_margins(10, 10, 10)
    for riga in nome_cognome.splitlines():
        pdf.multi_cell(0, 10, riga, align = 'C', )
    pdf.ln(5)

    # stampa data_nascita,luogo,ecc...
    pdf.set_margins(10, 10, 10)
    pdf.set_font("Helvetica", "I", size=12)
    pdf.set_text_color(255,250,250)
    pdf.cell(200, 5, txt=dati_utente, align = 'C', ln=True)
    pdf.ln(5)  # Aggiunge una nuova linea

def stampa_intestazione_pdf(pdf, image_path, header_text):
    # immagine ANPAL/UE in intestazione
    pdf.set_margins(10, 10, 10)
    pdf.image(image_path, x=10, y=8, w=190)  # Modifica i valori di x, y, e w secondo necessità
    pdf.ln(12)  # Spazio dopo l'immagine
    # Stampa dettaglio azienda
    azienda = info['azienda'] + " - P.IVA: " + info['partita_iva']
    settore_ateco = info['settore_ateco']
    ente_formativo = info['ente_formativo']
    ente_certificatore = info['ente_certificatore']
    stampa_dati_azienda(pdf, azienda, settore_ateco, ente_formativo, ente_certificatore)
    pdf.ln(20)  # Spazio dopo l'immagine
    # INTESTAZIONE DEL DOCUMENTO - SUBITO SOTTO IMMAGINE HEADER
    pdf.set_text_color(235, 9, 16)
    pdf.set_font("Helvetica", "B", size=36)
    pdf.cell(193, 7, txt=header_text, align = 'C', ln=True)
    pdf.ln(10)  # Spazio dopo header
    pdf.set_text_color(255,250,250)


# Funzione che stampa "per la partecipazione al corso di formazione generale e specifica" ed il relativo corso passato come parametro
def stampa_denominazione_percorso_sviluppo(pdf, nome_corso_da_stampare):
    pdf.ln(4)
    pdf.set_font("Helvetica", "BI", size=14)
    testo_header = "per la partecipazione al corso di formazione generale e specifica"
    pdf.cell(200, 7, txt=testo_header, align = 'C', ln=True)
    pdf.ln(4)
    pdf.set_text_color(235, 9, 16)
    pdf.set_font("Arial", "B", size=24)
    pdf.cell(200, 7, txt=nome_corso_da_stampare, align = 'C', ln=True)
    pdf.set_text_color(255,250,250)
    pdf.ln(8)

# Stampa testo in piu righe
def stampa_multicell(pdf, testo):
    pdf.add_font('DejaVu', '', './DejaVuSansCondensed.ttf', uni=True)
    pdf.set_font('DejaVu', '', 9)
    pdf.set_margins(10, 10, 10)
    for riga in testo.splitlines():
        pdf.multi_cell(0, 5, riga)

# stampa paragrafo generica con header char size = 12
def stampa_paragrafo(pdf,header_str,text_string):
    #stampa header
    pdf.set_font("Helvetica", "BI", size=12)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=header_str, align = 'C', ln=True)
    pdf.ln(3)
    # stampa testo
    stampa_multicell(pdf, text_string)

# stampa paragrafo secondario con header char size = 10
def stampa_paragrafo_secondario(pdf, header_str, text_string):
    # stampa header
    pdf.set_font("Helvetica", "B", size=12)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=header_str, align = 'C', ln=True)
    pdf.ln(1)
    # stampa testo
    stampa_multicell(pdf, text_string)

# STAMPA LA TABELLA PER LE FIRME DI ACCETTAZIONE DEI VARI ENTI
def stampa_firme_accettazione(pdf, data_emissione):
    
    # Stampa Luogo e Data
    pdf.set_margins(10, 10, 10)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(60, 10, 'Luogo e Data', 0, 0, 'C')
    pdf.set_font('Arial', 'I', 11)
    txt_date = 'Terni, lì ' + data_emissione
    pdf.text(25, pdf.get_y()+ 12, txt=txt_date)

    # Stampa Ente Formativo : OBM s.r.l.
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(60, 10, 'Il Docente', 0, 0, 'C')

    # Stampa Ente Certificatore
    pdf.cell(60, 10, 'Il Direttore del Corso', 0, 1, 'C')


# Funzione che salva il pdf e ci da in out i percorsi in cui sono stati salvati
def salvapdf(pdf, codice_fiscale, info, corso_id):
    # Salva il  pdf nella path /reports nominandolo come cf_siglacorso.pdf
    # Estrae la sigla del corso
    # Crea il percorso completo per il file PDF
    azienda = info['azienda']
    
    saved_file_name = f"{codice_fiscale}_{corso_id}.pdf"

    # Directory di test: output_path = os.path.join("prova_reports", azienda, saved_file_name)
    output_path = os.path.join("reports", azienda, saved_file_name)

    # Crea la cartella se non esiste
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    # Salva il PDF
    pdf.output(output_path, "F")
    print(f"Il file PDF è stato salvato in: {output_path}")


# funzione per convertire minuti in HH:MM
# serve per stampare il tempo dell'utente nel corso
def convert_minutes_to_hours_minutes(minutes):
    hours = int(minutes // 60)
    remaining_minutes = int(minutes % 60)
    return f"{hours:02d}:{remaining_minutes:02d}"

# funzione per convertire minuti in ore troncate
# serve per popolare la durata totale dei corsi
def convert_minutes_to_hours(minutes):
    hours = int(minutes // 60)
    return f"{hours}"

# Funzione per creare PDF per ogni codice fiscale
def crea_pdf(codice_fiscale, info):
    for corso in info['corsi']:
        # ----------------------------------------------------------------------------------------------------------------
        # INIZIO CREAZIONE DEL FOGLIO PDF
        # dopo la modifica del 12/07/2024 ad ogni funzione va passata la variabile PDF
        # ----------------------------------------------------------------------------------------------------------------
        pdf = FPDF()
        pdf.add_page()        
        # Imposta colore di riempimento RGB (esempio: azzurro chiaro)
        pdf.set_fill_color(36, 35, 35)
        pdf.set_text_color(255,250,250)
        
        # Disegna un rettangolo che copre tutta la pagina
        pdf.rect(0, 0, pdf.w, pdf.h, 'F')
        # immagine ANPAL/UE in intestazione e stampa "ALLEGATO 7 - ..." subito sotto l'immagine
        image_path = './intestazione.png'
        testo_header_1 = "ATTESTATO di FORMAZIONE"
        stampa_intestazione_pdf(pdf, image_path, testo_header_1)
        
        # stampa del blocco dati dell'azienda
        # [NON MODIFICARE] info['azienda'] ==> moodle_institution - nome dell'azienda richiedente
        # inserire le varie informazioni richieste
        
        # STAMPA DEL BLOCCO DATI UTENTE DA "Il presente Attestato viene conferito a:" FINO AI DATI ANAGRAFICI
        # in input passare in ordine le keys del dict info:
        # ['data_nascita'], ['comune_nascita'], ['cognome'], ['nome']
        # la variabile codice_fiscale
        
        blocco_dati_utente(pdf, info['data_nascita'], info['comune_nascita'], info['cognome'], info['nome'], codice_fiscale) # spaziatura verticale: pdf.ln(5)   
        pdf.ln()
        #########################################################################################################
        # Funzione che stampa i dettagli dell'utente relativi al percorso formativo: durata, ore totali e valutazione
        # passare in input il pdf, il dict del corso e la valutazione ottenuta al quiz 
        # momentaneamente la valutazione è fissa ma bisogna implementare la logica corretta    
        # valutazione = corso['valutazione_quiz']
        # DELETATA stampa_dettaglio_percorso_sviluppo(pdf, corso, valutazione)
        #########################################################################################################
        
        # IMPORTANTE: Dopo le modifiche del 20/05/2025, la chiave che lega master.xls a svil_comp.csv è:
        # il campo titolo_sessione_formativa di master è la chiave della colonna b di svil_comp.csv denominata perco_titolo
        # Quando dovremmo aggiungere una nuova tipologia di corso dovremmo aggiungere una riga al file svil_comp.csv
        # Nel file svil_comp bisogna compilare:
        # - la colonna perco_obiettivo con i dati relativi ai contenuti del corso
        # - la colonna perco_titolo con il titolo del corso
        # - la colonna perco_durata con il numero delle ore di durata del corso
        # Mentre nel file master.xlsx, quando si va ad aggiungere un utente che ha effettuato un corso bisogna
        # compilare come titolo_sessione_formativa lo stesso valore del campo perco_titolo del file svil_comp.csv
        # per assegnare all'utente un corso fatto e far stampare le denominazioni corrette sul PDF
        
        # DECODIFICA DELLA LISTA dettagli_percorso_formativo
        # [0] = perco_comp_id                   # [8] = comp_percorso
        # [1] = perco_titolo                    # [9] = comp_ref_areacompetenza
        # [2] = perco_obiettivo                 # [10] = comp_denominazione_competenza
        # [3] = perco_contenuto                 # [11] = comp_titolouc
        # [4] = perco_durata                    # [12] = comp_abilita
        # [5] = perco_tot_ore                   # [13] = comp_conoscenze
        # [6] = perco_destinatari
        # [7] = perco_mod_formativa
        
        denominazione_corso = corso['titolo_sessione_formativa']
        
        des_percorso_formativo = ''.join((list(denominazione_corso)))
        csv_filename = "./svil_comp.csv"
        dettagli_percorso_formativo = csv_to_list(csv_filename, des_percorso_formativo)
        
        if not dettagli_percorso_formativo or dettagli_percorso_formativo == None:
            print(f"Nessuna riga trovata con il titolo '{des_percorso_formativo}'.")   

        # Stampa NOME DEL CORSO = value scritto su titolo_sessione_formativa(master) e su perco_titolo(svil_comp)
        stampa_denominazione_percorso_sviluppo(pdf, dettagli_percorso_formativo[1]) # spaziatura verticale: pdf.ln(5)
        
        # Stampa CONTENUTI DEL CORSO
        # STAMPA [4] = perco_durata
        header_str = f"della durata complessiva di {dettagli_percorso_formativo[4]} ore, tenutosi in modalità FAD, con i seguenti contenuti:"
        # STAMPA [2] = perco_obiettivo
        text_string = ''.join((list(dettagli_percorso_formativo[2]))) 
        
        stampa_paragrafo(pdf, header_str, text_string)
        #########################################################################################################
        
        pdf.ln(20)
        
        # Stampa diciture per firme accettazione

        # Stampiamo la data di emissione sincronizzata ad ora
        data_certificazione = datetime.now()
        data_certificazione = data_certificazione.strftime("%d-%m-%Y")

        stampa_firme_accettazione(pdf, data_certificazione)
        #########################################################################################################
        
        corso = dettagli_percorso_formativo[0]
        # funzione di salvataggio del PDF
        salvapdf(pdf, codice_fiscale, info, corso) 

# ----------------------------------------------------------------------------------------------------------------    
# ------------------------------------- FINE FUNZIONI ------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------
#
#
# ----------------------------------------------------------------------------------------------------------------  
# ------------------------------- INIZIO GENERAZIONE DEI FILE PDF ------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------
# Usa il percorso assoluto del file Excel
#  file_path = './dati.xlsx'

# file di test: file_path = './master_ridotta.xlsx' #file di prova tabella master ridotta

# file_path = './master_ridotta.xlsx' #file di prova tabella master ridotta
file_path = './master_ridotta.xlsx' # file definitivo

# Leggi il file Excel
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Crea un dizionario per memorizzare i dati
data = {}
n_stampe = 0
# Itera attraverso le righe del foglio Excel
for row in sheet.iter_rows(min_row=2, values_only=True):  # Salta l'intestazione
    moodle_institution, cognome, nome, codice_fiscale, data_di_nascita, comune_di_nascita, titolo_azione_con_id, titolo_sessione_formativa, Totaleorecorso, durata_min, stato_corso, valutazione_quiz, partita_iva, settore_ateco, ente_formativo, ente_certificatore  = row

    if codice_fiscale is None:
        continue  # Salta la riga se il codice fiscale è None

    if codice_fiscale not in data:
        data[codice_fiscale] = {
            'azienda': moodle_institution,
            'nome': nome,
            'cognome': cognome,
            'data_nascita' : data_di_nascita.strftime("%d/%m/%Y"),
            'comune_nascita': comune_di_nascita,
            'partita_iva': partita_iva, 
            'settore_ateco': settore_ateco, 
            'ente_formativo': ente_formativo, 
            'ente_certificatore': ente_certificatore,
            'corsi': []
        }
    
    data[codice_fiscale]['corsi'].append({
        'titolo_azione_con_id': titolo_azione_con_id,
        'titolo_sessione_formativa': titolo_sessione_formativa,
        'Totaleorecorso': Totaleorecorso,
        'durata_min': durata_min,
        'stato_corso': stato_corso,
        'valutazione_quiz': valutazione_quiz
    })
    n_stampe = n_stampe + 1

# Crea PDF per ogni codice fiscale
for codice_fiscale, info in data.items():
    crea_pdf(codice_fiscale, info)

print("N. totale stampe: "+str(n_stampe))