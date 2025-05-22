import openpyxl

def estrai_data_nascita(codice_fiscale):
    mesi_codice = 'ABCDEHLMPRST'  # Mesi: A=Gen, B=Feb, ..., T=Dic
    anno = int(codice_fiscale[6:8])
    mese = mesi_codice.index(codice_fiscale[8]) + 1
    giorno = int(codice_fiscale[9:11])
    if giorno > 40:
        giorno -= 40
    if anno < 40:
        anno += 2000
    else:
        anno += 1900
    return f"{giorno}/{mese}/{anno}"

def xls_to_dict(xls_filename, id_corso):
    try:
        wb = openpyxl.load_workbook(xls_filename, data_only=True)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == str(id_corso).strip():
                return {headers[i]: row[i] for i in range(len(headers))}
    except Exception as e:
        print(f"Errore nella lettura del file {xls_filename}: {e}")
    return None

def carica_db_corsi(path_db_corsi):
    """
    Carica db_corsi.xlsx in un dizionario con chiave 'odoo_id'.
    Restituisce dict[id_corso] = {'odoo_id', 'nome_corso', 'contenuti_corso', 'durata_corso','save_path'}
    """
    try:
        wb = openpyxl.load_workbook(path_db_corsi, data_only=True)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        expected_headers = ['odoo_id', 'nome_corso', 'contenuti_corso', 'durata_corso','save_path']
        for eh in expected_headers:
            if eh not in headers:
                raise ValueError(f"Header '{eh}' mancante in db_corsi.xlsx")

        corsi = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            id_corso = str(row[headers.index('odoo_id')]).strip()
            corsi[id_corso] = {
                "nome_corso": row[headers.index('nome_corso')],
                "contenuti_corso": row[headers.index('contenuti_corso')],
                "durata_corso": row[headers.index('durata_corso')],
                "save_path": row[headers.index('save_path')]
            }
        return corsi
    except Exception as e:
        print(f"Errore durante il caricamento di db_corsi: {e}")
        return {}

def carica_aziende(path_aziende):
    """
    Carica aziende.xlsx in un dizionario indicizzato per id_azienda o ragione_sociale.
    """
    try:
        wb = openpyxl.load_workbook(path_aziende, data_only=True)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]

        aziende = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Usa id_azienda se presente (prima colonna), altrimenti ragione_sociale (seconda)
            key = row[0] if row[0] else row[1]
            if key:
                aziende[str(key).strip()] = {headers[i]: row[i] for i in range(len(headers))}
        return aziende
    except Exception as e:
        print(f"Errore durante il caricamento di aziende: {e}")
        return {}

def carica_stampa_pdf(path_stampa_pdf):
    wb = openpyxl.load_workbook(path_stampa_pdf, data_only=True)
    sheet = wb.active
    headers = [cell.value for cell in sheet[1]]

    idx_cf = headers.index("codice_fiscale")
    idx_nome = headers.index("nominativo")
    idx_comune = headers.index("comune")  # <-- qui correggi
    idx_id_azienda = headers.index("id_azienda") if "id_azienda" in headers else None
    idx_id_corso = headers.index("id_corso")
    idx_docente = headers.index("docente")

    data = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        cf = row[idx_cf]
        if not cf:
            continue
        if cf not in data:
            data[cf] = {
                "nominativo": row[idx_nome],
                "data_nascita": estrai_data_nascita(cf),
                "luogo_nascita": row[idx_comune],  # <-- qui assegni comune a luogo_nascita
                "id_azienda": row[idx_id_azienda] if idx_id_azienda is not None else None,
                "corsi": []
            }
        corso = {
            "id_corso": row[idx_id_corso],
            "docente": row[idx_docente]
        }
        data[cf]["corsi"].append(corso)
    return data
