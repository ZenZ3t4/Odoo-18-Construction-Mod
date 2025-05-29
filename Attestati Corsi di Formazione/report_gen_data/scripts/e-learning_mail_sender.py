import os
import smtplib
import ssl
from openpyxl import load_workbook
from email.message import EmailMessage
import logging


# === CONFIGURAZIONE ===
EMAIL_SENDER = "contatorecoopbastia@gmail.com"
APP_PASSWORD = "yszb mezf uter fzub"
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 465
EXCEL_PATH = "test_sender_mail.xlsx"
PDF_DIR = "./E-Learning Odoo Reports"

# === Logger ===
log_dir = './logs/e-mail sender'
os.makedirs(log_dir, exist_ok=True)
log_file_path = os.path.join(log_dir, 'mail_log.log')
logging.basicConfig(
    filename=log_file_path,
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level=logging.INFO
)

# === FUNZIONE: Lettura dati Excel ===
def leggi_contatti(file_path):
    contatti = []
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Trova gli indici delle colonne giuste
    intestazioni = [cell.value for cell in sheet[1]]
    idx_nominativo = intestazioni.index("nominativo")
    idx_cf = intestazioni.index("codice_fiscale")
    idx_email = intestazioni.index("e_mail")

    # Leggi i dati riga per riga
    for row in sheet.iter_rows(min_row=2, values_only=True):
        nominativo = row[idx_nominativo]
        cf = row[idx_cf]
        email = row[idx_email]
        contatti.append({
            "nominativo": nominativo,
            "cf": cf,
            "email": email
        })
    return contatti


# === FUNZIONE: Ricerca file PDF per codice fiscale (ricorsiva) ===
def trova_pdf_per_cf(cartella, codice_fiscale):
    pdf_trovati = []
    for root, dirs, files in os.walk(cartella):
        for filename in files:
            if codice_fiscale.upper() in filename.upper() and filename.lower().endswith(".pdf"):
                pdf_trovati.append(os.path.join(root, filename))
    return pdf_trovati

# === FUNZIONE: Composizione e invio email ===
def invia_email(destinatario, nominativo, allegati):
    msg = EmailMessage()
    msg["Subject"] = "Attestati Formazione"
    msg["From"] = EMAIL_SENDER
    msg["To"] = destinatario
    msg.set_content(f"Gentile dipendente,\n\nin allegato a questa mail trova le certificazioni aggiornate in conformità alle normative vigenti, relative ai corsi da lei svolti attraverso il nostro portale. La invitiamo a prenderne visione e a considerarle come attestati ufficiali della formazione da lei completata. \n\nCordiali Saluti,\nAuthentica S.p.A.")
 
    for filepath in allegati:
        with open(filepath, "rb") as f:
            file_data = f.read()
            file_name = os.path.basename(filepath)
            msg.add_attachment(file_data, maintype="application", subtype="pdf", filename=file_name)

    # Invio email
    try:
        context = ssl.create_default_context()
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, context=context) as server:
            server.login(EMAIL_SENDER, APP_PASSWORD)
            server.send_message(msg)
        print(f"[OK] Email inviata a {destinatario}")
        logging.info(f"[OK] Email inviata a {destinatario}")
    except Exception as e:
        print(f"[ERRORE] Email a {destinatario} fallita: {e}")
        logging.error(f"[ERRORE] Email a {destinatario} fallita: {e}")

# === FUNZIONE PRINCIPALE ===
def main():
    contatti = leggi_contatti(EXCEL_PATH)
    for contatto in contatti:
        nominativo = contatto["nominativo"]
        cf = contatto["cf"]
        email = contatto["email"]
        allegati = trova_pdf_per_cf(PDF_DIR, cf)

        if not allegati:
            print(f"[AVVISO] Nessun PDF trovato per {nominativo} ({cf}) - Email non inviata")
            logging.error(f"[AVVISO] Nessun PDF trovato per {nominativo} ({cf}) - Email non inviata")
            continue

        invia_email(email, nominativo, allegati)

# === AVVIO SCRIPT ===
if __name__ == "__main__":
    main()
