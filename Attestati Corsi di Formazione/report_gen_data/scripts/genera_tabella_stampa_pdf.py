import openpyxl
import os
import uuid
import logging
from datetime import datetime

# === CONFIGURAZIONE LOG ===
log_dir = './report_gen_data/logs/'
os.makedirs(log_dir, exist_ok=True)
log_file_path = os.path.join(log_dir, 'merge_report.log')
logging.basicConfig(
    filename=log_file_path,
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO
)

def trova_file_odoo(folder, azienda):
    for filename in os.listdir(folder):
        if (filename.endswith(".xls") or filename.endswith(".xlsx")) and azienda in filename:
            return os.path.join(folder, filename)
    return None

# === PERCORSI FILE ===
file_anagrafica = './report_gen_data/db/db_anagrafica_new_reports.xlsx'
file_output = './report_gen_data/db/stampa_pdf.xlsx'
file_merge_error_xlsx = './report_gen_data/logs/merge_errors.xlsx'
folder_session_output = './report_gen_data/db/xls_stampa'
os.makedirs(folder_session_output, exist_ok=True)

try:
    input_file_id_odoo = input("Inserisci identificativo file odoo: ")
    folder_path = "./report_gen_data/odoo_data"
    file_lista_odoo = trova_file_odoo(folder_path, input_file_id_odoo)

    if file_lista_odoo is None:
        msg = f"Nessun file Excel trovato in '{folder_path}' contenente '{input_file_id_odoo}' nel nome."
        logging.error(msg)
        raise FileNotFoundError(msg)

    print(f"File trovato: {file_lista_odoo}")
    logging.info(f"File Odoo trovato: {file_lista_odoo}")

    docente = input("Inserisci il nome del docente da associare a tutti i record: ").strip()
    logging.info(f"Docente selezionato: {docente}")

    # === CARICAMENTO MASTER ===
    wb_master = openpyxl.load_workbook(file_anagrafica)
    sheet_master = wb_master.active
    headers_master = [cell.value for cell in sheet_master[1]]

    # === CARICAMENTO ODOO ===
    wb_odoo = openpyxl.load_workbook(file_lista_odoo)
    sheet_odoo = wb_odoo.active
    headers_odoo = [cell.value for cell in sheet_odoo[1]]

    idx_email_master = headers_master.index("e_mail")
    colonne_output = headers_master + ["id_corso", "docente", "record_id"]
    idx_codice_fiscale = headers_master.index("codice_fiscale")

    # === Dizionario anagrafico ===
    diz_anagrafica = {}
    for row in sheet_master.iter_rows(min_row=2, values_only=True):
        email = row[idx_email_master]
        if email:
            diz_anagrafica[email.strip().lower()] = row

    # === APERTURA O CREAZIONE file_output ===
    if os.path.exists(file_output):
        wb_output = openpyxl.load_workbook(file_output)
        sheet_output = wb_output.active
    else:
        wb_output = openpyxl.Workbook()
        sheet_output = wb_output.active
        sheet_output.append(colonne_output)

    # === NUOVO FILE PER SESSIONE CORRENTE ===
    now = datetime.now()
    timestamp_str = now.strftime("%Y%m%d_%H%M")
    nome_corso_sample = ""
    wb_session = openpyxl.Workbook()
    sheet_session = wb_session.active
    sheet_session.append(colonne_output)

     # === OUTPUT ERRORI ===
    folder_error_output = './report_gen_data/logs/errors'
    os.makedirs(folder_error_output, exist_ok=True)

    wb_error = openpyxl.Workbook()
    sheet_error = wb_error.active
    sheet_error.append(["E-mail", "Corso", "Motivo"])
    
    # Controllo duplicati
    existing_keys = set()
    if os.path.exists(file_output):
        wb_output = openpyxl.load_workbook(file_output)
        sheet_output = wb_output.active
        existing_headers = [cell.value for cell in sheet_output[1]]
        try:
            idx_cf_output = existing_headers.index("codice_fiscale")
            idx_corso_output = existing_headers.index("id_corso")
            for row in sheet_output.iter_rows(min_row=2, values_only=True):
                if row[idx_cf_output] and row[idx_corso_output]:
                    key = (row[idx_cf_output], row[idx_corso_output])
                    existing_keys.add(key)
        except ValueError:
            pass  # file vuoto o non ancora strutturato
    else:
        wb_output = openpyxl.Workbook()
        sheet_output = wb_output.active
        sheet_output.append(colonne_output)

    righe_scritti = 0
    righe_mancanti = 0

    for row in sheet_odoo.iter_rows(min_row=2, values_only=True):
        try:
            email_odoo = row[headers_odoo.index("E-mail")].strip().lower()
            codice_corso = row[headers_odoo.index("Corso")]
            nome_corso_sample = codice_corso  # usato per nome file

            if email_odoo in diz_anagrafica:
                dati = list(diz_anagrafica[email_odoo])  # <-- QUI
                
                codice_fiscale = dati[idx_codice_fiscale]  # usato per controllare duplicati
                if (codice_fiscale, codice_corso) in existing_keys:
                    logging.info(f"Duplicato evitato per {codice_fiscale} - Corso: {codice_corso}")
                    continue

                record_id = str(uuid.uuid4())
                dati.append(codice_corso)
                dati.append(docente)
                dati.append(record_id)
                sheet_output.append(dati)
                sheet_session.append(dati)
                existing_keys.add((codice_fiscale, codice_corso))
                righe_scritti += 1
            else:
                msg = f"[!] Email non trovata nel master: {email_odoo}"
                print(msg)
                logging.warning(msg)
                sheet_error.append([email_odoo, codice_corso, "Email non presente in db_anagrafica"])
                righe_mancanti += 1

        except Exception as e:
            logging.error(f"Errore durante l'elaborazione della riga Odoo: {row} -> {e}")
            sheet_error.append(["ERRORE", "ERRORE", f"{str(e)}"])

    # === Salvataggio file cumulativo
    wb_output.save(file_output)

    # === Salvataggio file sessione
    nome_file_sessione = f"stampa_{nome_corso_sample}_{timestamp_str}.xlsx".replace(" ", "_")
    path_file_sessione = os.path.join(folder_session_output, nome_file_sessione)
    wb_session.save(path_file_sessione)

    # === Salvataggio file errori
    nome_file_errori = f"errors_{nome_corso_sample}_{timestamp_str}.xlsx".replace(" ", "_")
    path_file_errori = os.path.join(folder_error_output, nome_file_errori)
    wb_error.save(path_file_errori)

    print(f"\n✅ File 'stampa_pdf.xlsx' aggiornato.")
    print(f"📁 File sessione salvato in: {path_file_sessione}")
    print(f"⚠️ File errori salvato in: {path_file_errori}")
    logging.info(f"PDF: {righe_scritti} righe scritte, {righe_mancanti} mancanti.")
    logging.info(f"Errori salvati in: {path_file_errori}")
    logging.info("Script completato con successo.")

except Exception as e:
    logging.critical(f"Errore critico durante l'esecuzione dello script: {e}")
    raise

