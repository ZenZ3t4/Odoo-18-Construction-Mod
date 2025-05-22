from funzioni_reports import *
import openpyxl
from fpdf import FPDF
import os

# #####################################################
# -----------------------------------------
# FUNZIONI UTILIZZATE NEL CODICE
# -----------------------------------------

# STAMPA BLOCCO DATI AZIENDA
def stampa_dati_azienda(pdf, azienda, settore_ateco, ente_formativo, ente_certificatore):
    # BLOCCO DATI AZIENDA
    pdf.set_font("Arial", "B", size=12)
    azienda_header = "Dati identificativi dell'azienda"
    pdf.cell(200, 5, txt=azienda_header, ln=True)

    if len(settore_ateco) > 80:
        settore_ateco = settore_ateco[:80] + '...'
    
    # Dati dinamici azienda
    pdf.set_font("Helvetica", size=10)
    azienda_row_1 = f"Azienda: {azienda}"
    azienda_row_2 = f"Settore ateco: {settore_ateco}"
    azienda_row_3 = f"Ente formativo: {ente_formativo}"
    azienda_row_4 = f"Ente certificatore: {ente_certificatore}"
    pdf.cell(200, 5, txt=azienda_row_1, ln=True)
    pdf.cell(200, 5, txt=azienda_row_2, ln=True)
    pdf.cell(200, 5, txt=azienda_row_3, ln=True)
    pdf.cell(200, 5, txt=azienda_row_4, ln=True)
    pdf.ln(10)  # Spazio dopo blocco dati azienda
    #--------------------------------------------------------------------------------


# STAMPA BLOCCO DATI UTENTE
def blocco_dati_utente(pdf, data_nascita, luogo_nascita, cognome, nome, codice_fiscale):
    # BLOCCO DATI UTENTE
    # header in grassetto
    pdf.set_font("Arial", "B", size=16)
    pdf.set_margins(20, 15, 20)
    testo_rilasciato_a = "RILASCIATO A"
    pdf.cell(200, 10, txt=testo_rilasciato_a, align = 'C', ln=True)
    pdf.ln(5)
    # blocco dati dinamici
    dati_utente = f"nato a {luogo_nascita} il {data_nascita}  - C.F. {codice_fiscale}"
    # dati anagrafici del candidato
    nome_cognome = nome + ' ' + cognome
    pdf.set_font("Helvetica", 'B', size=22)
    pdf.set_margins(20, 15, 20)
    for riga in nome_cognome.splitlines():
        pdf.multi_cell(0, 10, riga, align = 'C', )
    pdf.ln(7)

    # stampa data_nascita,luogo,ecc...
    pdf.set_margins(10, 10, 10)
    pdf.set_font("Helvetica", size=11)
    pdf.cell(200, 5, txt=dati_utente, ln=True)
    pdf.ln(10)  # Aggiunge una nuova linea


# STAMPA IMMAGINE ANPAL/UE in intestazione
def stampa_intestazione_pdf(pdf, image_path, header_text):
    # immagine ANPAL/UE in intestazione
    pdf.set_margins(10, 10, 10)
    pdf.image(image_path, x=10, y=8, w=190)  # Modifica i valori di x, y, e w secondo necessità
    pdf.ln(20)  # Spazio dopo l'immagine
    
    # INTESTAZIONE DEL DOCUMENTO - SUBITO SOTTO IMMAGINE HEADER
    pdf.set_font("Arial", "B", size=12)
    pdf.cell(193, 7, txt=header_text, align = 'C', ln=True)
    # stampiamo in due righe altrimenti non ci entra in larghezza
    # pdf.set_font("Arial", "B", size=13)
    # testo_header_1 = " "
    # pdf.cell(200, 7, txt=testo_header_1, align = 'C', ln=True)
    pdf.ln(10)  # Spazio dopo header

# STAMPA IL PARAGRAFO "ATTESTATO FINALE DI MESSA IN TRASPARENZA DELLE COMPETENZE"
def stampa_header_meta_pagina(pdf):
    # header metà attestato
    # header in grassetto
    pdf.set_font("Arial", "B", size=16)
    testo_header_2 = "ATTESTATO FINALE DI MESSA IN TRASPARENZA DELLE "
    pdf.cell(200, 7, txt=testo_header_2, align = 'C', ln=True)
    testo_header_2 = "COMPETENZE"
    pdf.cell(200, 7, txt=testo_header_2, align = 'C', ln=True)
    
    # header in corsivo
    pdf.set_font("Arial", "I", size=11)
    testo_subheader = "Relativa al percorso formativo realizzato nell'ambito di interventi finanziati dal Fondo Nuove Competenze"
    pdf.cell(193, 7, txt=testo_subheader, align = 'C', ln=True)
    pdf.ln(5)  # Aggiunge una nuova linea


# Funzione che stampa "DENOMINAZIONE DEL PERCORSO..." ed il relativo corso passato come parametro
def stampa_denominazione_percorso_sviluppo(pdf, denominazione_corso):
    pdf.set_font("Arial", "B", size=14)
    testo_header = "DENOMINAZIONE DEL PERCORSO DI SVILUPPO DELLE COMPETENZE:"
    pdf.cell(200, 7, txt=testo_header, align = 'C', ln=True)
    pdf.ln(2)
    pdf.cell(200, 7, txt=denominazione_corso, align = 'C', ln=True)
    pdf.ln(8)


# Funzione che stampa i dettagli dell'utente relativi al percorso formativo: durata, ore totali e valutazione
def stampa_dettaglio_percorso_sviluppo(pdf, corso, valutazione):
    # creazione stringhe
    esito_corso = corso['stato_corso']
    durata_ore = convert_minutes_to_hours_minutes(corso['Totaleorecorso'])
    totale_ore_corso = convert_minutes_to_hours(corso['durata_min'])
    testo_intro_corsi = f"{corso['titolo_azione_con_id']}"
    testo_corsi_2 = f"- Durata del percorso di {totale_ore_corso} ore di cui {totale_ore_corso} ore di FAD asincrona. Ore effettive svolte: {durata_ore} ore."
    
    #stampa stringhe
    pdf.set_font("Helvetica", "B", size=13)
    pdf.cell(200, 5, txt=testo_intro_corsi, ln=True)
    pdf.ln(2)
    pdf.set_font("Helvetica", size=12)
    pdf.cell(200, 5, txt=testo_corsi_2, ln=True)
    pdf.ln(1)

    # se il corso è incompleto, la stampa diventa in grassetto
    if esito_corso == 'completo':
        testo_corsi_3 = f"- Corso {esito_corso}  -  Valutazione finale: {valutazione}  -  valutazione effettuata tramite quiz."
    else:
        pdf.set_font("Helvetica", "B", size=12)
        testo_corsi_3 = f"- Corso {esito_corso}"

    pdf.cell(200, 5, txt=testo_corsi_3, ln=True)
    pdf.ln(5)
    pdf.set_font("Helvetica", size=12)


# STAMPA TESTO IN PIU RIGHE
def stampa_multicell(pdf, testo):
    pdf.add_font('DejaVu', '', './DejaVuSansCondensed.ttf', uni=True)
    pdf.set_font('DejaVu', '', 10)
    pdf.set_margins(20, 15, 20)
    for riga in testo.splitlines():
        pdf.multi_cell(0, 5, riga)

# stampa paragrafo generica con header char size 12
def stampa_paragrafo(pdf,header_str,text_string):
    #stampa header
    pdf.set_font("Helvetica", "B", size=12)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=header_str, align = 'C', ln=True)
    pdf.ln(1)
    # stampa testo
    stampa_multicell(pdf, text_string)

# stampa paragrafo secondario con header char size = 10
def stampa_paragrafo_secondario(pdf, header_str, text_string):
    # stampa header
    pdf.set_font("Helvetica", "B", size=10)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=header_str, align = 'C', ln=True)
    pdf.ln(1)
    # stampa testo
    stampa_multicell(pdf, text_string)

# STAMPA LA TABELLA PER LE FIRME DI ACCETTAZIONE DEI VARI ENTI
def stampa_firme_accettazione(pdf, ente_formativo_img_path, ente_formativo, ente_certificatore, ente_certificatore_img_path):
    # Controllo che l'immagine dell'ente certificatore esista
    # N.B: L'immagine dell'ente certificatore deve essere nominata esattamente come il nome dell'ente
    # Esempio per THE LANGUAGE CENTER S.R.L.: file_name = THE LANGUAGE CENTER S.R.L..png
    if ente_certificatore_img_path == "./.png":
            ente_certificatore_img_path = None

    # Stampa Luogo e Data
    pdf.set_margins(10, 10, 10)
    pdf.ln(35)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(60, 10, 'Luogo e Data', 0, 0, 'C')
    pdf.set_font('Arial', 'I', 11)
    txt_date = 'Terni, lì 26/02/2024'
    pdf.text(25, pdf.get_y()+ 12, txt=txt_date)

    # Stampa Ente Formativo : OBM s.r.l.
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(60, 10, 'Per '+ ente_formativo, 0, 0, 'C')
    pdf.image(ente_formativo_img_path, 85, pdf.get_y() + 12, 30)

    # Stampa Ente Certificatore
    pdf.cell(60, 10, 'Per '+ ente_certificatore, 0, 1, 'C')
    if ente_certificatore_img_path != None:
        pdf.image(ente_certificatore_img_path, 145, pdf.get_y()+ 12, 30)


# Funzione che salva il pdf e ci da in out i percorsi in cui sono stati salvati
def salvapdf(pdf, codice_fiscale, info, denominazione_corso, corso):
    # Salva il  pdf nella path /reports nominandolo come cf_siglacorso.pdf
    # Estrae la sigla del corso
    sigla_corso = list(denominazione_corso)
    sigla_corso = str(sigla_corso[6])
    
    # Crea il percorso completo per il file PDF
    # Se il corso è incompleto stampiamo una dicitura differente sul nome del pdf
    if corso['stato_corso'] == 'incompleto':
        saved_file_name = f"{codice_fiscale}_{sigla_corso}_inc.pdf"
    else:
        saved_file_name = f"{codice_fiscale}_{sigla_corso}.pdf"

    # Directory di test: output_path = os.path.join("prova_reports", saved_file_name)
    output_path = os.path.join(f"single_user_reports/{codice_fiscale}", saved_file_name)
    
    # Crea la cartella se non esiste
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    # Salva il PDF
    pdf.output(output_path, "F")
    print(f"Il file PDF è stato salvato in: {output_path}")


# funzione per convertire minuti in HH:MM
# serve per stampare il tempo dell'utente nel corso
def convert_minutes_to_hours_minutes(minutes):
    hours = int(minutes // 60)
    remaining_minutes = int(minutes % 60)
    return f"{hours:02d}:{remaining_minutes:02d}"

# funzione per convertire minuti in ore troncate
# serve per popolare la durata totale dei corsi
def convert_minutes_to_hours(minutes):
    hours = int(minutes // 60)
    return f"{hours}"

# Funzione per creare PDF per ogni codice fiscale
def crea_pdf(codice_fiscale, info):
    for corso in info['corsi']:
        # ----------------------------------------------------------------------------------------------------------------
        # INIZIO CREAZIONE DEL FOGLIO PDF
        # dopo la modifica del 12/07/2024 ad ogni funzione va passata la variabile PDF
        # ----------------------------------------------------------------------------------------------------------------
        pdf = FPDF()
        pdf.add_page()        
        
        # immagine ANPAL/UE in intestazione e stampa "ALLEGATO 7 - ..." subito sotto l'immagine
        image_path = './intestazione.png'
        testo_header_1 = "ALLEGATO 7 - ATTESTATO FINALE DI MESSA IN TRASPARENZA DELLE COMPETENZE"
        stampa_intestazione_pdf(pdf, image_path, testo_header_1)
        
        # stampa del blocco dati dell'azienda
        # [NON MODIFICARE] info['azienda'] ==> moodle_institution - nome dell'azienda richiedente
        # inserire le varie informazioni richieste
        azienda = info['azienda'] + " - P.IVA: " + info['partita_iva']
        settore_ateco = info['settore_ateco']
        ente_formativo = info['ente_formativo']
        ente_certificatore = info['ente_certificatore']
        stampa_dati_azienda(pdf, azienda, settore_ateco, ente_formativo, ente_certificatore) # spaziatura: pdf.ln(10)

        # stampa "ATTESTATO FINALE DI ..." a metà pagina - NON RICHIEDE PARAMETRI, E' UNA STRINGA FISSA
        stampa_header_meta_pagina(pdf)
        
        # STAMPA DEL BLOCCO DATI UTENTE DA "RILASCIATO A" FINO AI DATI ANAGRAFICI
        # in input passare in ordine le keys del dict info:
        # ['data_nascita'], ['comune_nascita'], ['cognome'], ['nome']
        # la variabile codice_fiscale
        blocco_dati_utente(pdf, info['data_nascita'], info['comune_nascita'], info['cognome'], info['nome'], codice_fiscale) # spaziatura verticale: pdf.ln(5)   
        pdf.ln()
        
        # Funzione che stampa "DENOMINAZIONE DEL PERCORSO..." ed il relativo corso passato come parametro      
        denominazione_corso = corso['titolo_sessione_formativa']
        stampa_denominazione_percorso_sviluppo(pdf, denominazione_corso) # spaziatura verticale: pdf.ln(5)

        # Funzione che stampa i dettagli dell'utente relativi al percorso formativo: durata, ore totali e valutazione
        # passare in input il pdf, il dict del corso e la valutazione ottenuta al quiz 
        # momentaneamente la valutazione è fissa ma bisogna implementare la logica corretta    
        valutazione = corso['valutazione_quiz']
        stampa_dettaglio_percorso_sviluppo(pdf, corso, valutazione)
        
        #########################################################################################################
        # dettagli_percorso_formativo è una lista che contiene i dettagli delle varie competenze acquisite nei corsi
        des_percorso_formativo = ''.join((list(denominazione_corso)[9:]))
        csv_filename = "svil_comp.csv"
        dettagli_percorso_formativo = csv_to_list(csv_filename, des_percorso_formativo)
        
        if not dettagli_percorso_formativo:
            print(f"Nessuna riga trovata con il titolo '{des_percorso_formativo}'.")     
            
        # DECODIFICA DELLA LISTA dettagli_percorso_formativo
        # [0] = perco_comp_id                   # [8] = comp_percorso
        # [1] = perco_titolo                    # [9] = comp_ref_areacompetenza
        # [2] = perco_obiettivo                 # [10] = comp_denominazione_competenza
        # [3] = perco_contenuto                 # [11] = comp_titolouc
        # [4] = perco_durata                    # [12] = comp_abilita
        # [5] = perco_tot_ore                   # [13] = comp_conoscenze
        # [6] = perco_destinatari
        # [7] = perco_mod_formativa
            
        # Stampa OBIETTIVO DEL PERCORSO FORMATIVO
        pdf.ln(7)
        header_str = "APPRENDIMENTI CONSEGUITI"
        text_string = ''.join((list(dettagli_percorso_formativo[2]))) # STAMPA [2] = perco_obiettivo
        stampa_paragrafo(pdf, header_str, text_string)
        pdf.ln(2)

        # Stampa ADA
        header_str = "Denominazione Area di attivita' (ADA) dell' Atlante del lavoro e delle qualificazioni: "
        text_string = ''.join((list(dettagli_percorso_formativo[9]))) # STAMPA [9] = comp_ref_areacompetenza
        stampa_paragrafo_secondario(pdf, header_str, text_string)
        
        # AGGIUNGIAMO UNA SECONDA PAGINA E L'INTESTAZIONE
        pdf.add_page()

        # immagine ANPAL/UE in intestazione e stampa "ALLEGATO 7 - ..." subito sotto l'immagine
        testo_header_1 = "ALLEGATO 7 - ATTESTATO FINALE DI MESSA IN TRASPARENZA DELLE COMPETENZE"
        stampa_intestazione_pdf(pdf, image_path, testo_header_1)
        pdf.ln(8)
        
        # STAMPA ATTIVITA E RISULTATI ATTESI
        testo_header_2 = "ATTIVITA' E RISULTATI ATTESI"
        text_string = ''.join((list(dettagli_percorso_formativo[11]))) # [11] = comp_titolouc
        stampa_paragrafo(pdf, testo_header_2, text_string)
        text_string = ''.join((list(dettagli_percorso_formativo[10]))) # [10] = comp_denominazione_competenza
        stampa_multicell(pdf, text_string)
        pdf.ln(7)

        # STAMPA ABILITA
        testo_header_2 = "ABILITA'"
        text_string = ''.join((list(dettagli_percorso_formativo[12]))) # [12] = comp_abilita
        stampa_paragrafo(pdf, testo_header_2, text_string)
        pdf.ln(7)

        # stampa conoscenze
        testo_header_2 = "CONOSCENZE"
        text_string = ''.join((list(dettagli_percorso_formativo[13]))) # [13] = comp_conoscenze
        stampa_paragrafo(pdf, testo_header_2, text_string)

        # Stampa diciture per firme accettazione
        # Creazione della tabella

        pdf.ln(13)
        ente_formativo_img_path = "./obm_timbro.png"
        ente_certificatore_img_path = f"./{ente_certificatore}.png"
        # print(ente_certificatore_img_path)
        stampa_firme_accettazione(pdf, ente_formativo_img_path, ente_formativo, ente_certificatore, ente_certificatore_img_path)
        #########################################################################################################
        
        # funzione di salvataggio del PDF
        salvapdf(pdf, codice_fiscale, info, denominazione_corso, corso) 
# ----------------------------------------------------------------------------------------------------------------    
# ------------------------------------- FINE FUNZIONI ------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------
#
#
# ----------------------------------------------------------------------------------------------------------------  
# ------------------------------- INIZIO GENERAZIONE DEI FILE PDF ------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------
# Usa il percorso assoluto del file Excel
#  file_path = './dati.xlsx'

# file di test: file_path = './master_ridotta.xlsx' #file di prova tabella master ridotta

# file_path = './master_ridotta.xlsx' #file di prova tabella master ridotta
file_path = './master.xlsx' # file definitivo

# Leggi il file Excel
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Crea PDF dato in input un codice_fiscale
cf_utente_input = str(input("inserisci codice fiscale: ")).upper()
# CF di test: ZZINRE02H05B963E

# Crea un dizionario per memorizzare i dati
data = {}
n_stampe = 0
# Itera attraverso le righe del foglio Excel
for row in sheet.iter_rows(min_row=2, values_only=True):  # Salta l'intestazione
    
    moodle_institution, cognome, nome, codice_fiscale, data_di_nascita, comune_di_nascita, titolo_azione_con_id, titolo_sessione_formativa, Totaleorecorso, durata_min, stato_corso, valutazione_quiz, partita_iva, settore_ateco, ente_formativo, ente_certificatore  = row
    if codice_fiscale == cf_utente_input :
        if codice_fiscale is None:
            continue  # Salta la riga se il codice fiscale è None

        if codice_fiscale not in data:
            data[codice_fiscale] = {
                'azienda': moodle_institution,
                'nome': nome,
                'cognome': cognome,
                'data_nascita' : data_di_nascita.strftime("%d/%m/%Y"),
                'comune_nascita': comune_di_nascita,
                'partita_iva': partita_iva, 
                'settore_ateco': settore_ateco, 
                'ente_formativo': ente_formativo, 
                'ente_certificatore': ente_certificatore,
                'corsi': []
            }
        
        data[codice_fiscale]['corsi'].append({
            'titolo_azione_con_id': titolo_azione_con_id,
            'titolo_sessione_formativa': titolo_sessione_formativa,
            'Totaleorecorso': Totaleorecorso,
            'durata_min': durata_min,
            'stato_corso': stato_corso,
            'valutazione_quiz': valutazione_quiz
        })
        n_stampe = n_stampe + 1
        



for codice_fiscale, info in data.items():
    if codice_fiscale == cf_utente_input:
        crea_pdf(cf_utente_input, info)

print("N. totale stampe: "+str(n_stampe))