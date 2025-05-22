from funzioni_reports import *
import openpyxl
from fpdf import FPDF
from datetime import datetime
import os

# #####################################################
# -----------------------------------------
# FUNZIONI UTILIZZATE NEL CODICE
# -----------------------------------------
def stampa_dati_azienda(pdf, azienda, settore_ateco, ente_formativo, ente_certificatore):
    # BLOCCO DATI AZIENDA
    if len(settore_ateco) > 80:
        settore_ateco = settore_ateco[:80] + '...'

    # Dati dinamici azienda
    pdf.set_font("Helvetica", "I", size=11)
    #azienda_row_1 = f"{azienda} - Settore ateco: {settore_ateco}"
    azienda_row_1 = f"collaboratore dell'azienda {azienda}"
    pdf.cell(200, 5, txt=azienda_row_1, align = 'C', ln=True)
    #--------------------------------------------------------------------------------
    
def blocco_dati_utente(pdf, data_nascita, luogo_nascita, cognome, nome, codice_fiscale):
    # BLOCCO DATI UTENTE
    # header in grassetto
    pdf.set_font("Helvetica", "BI", size=14)
    pdf.set_margins(10, 10, 10)
    testo_rilasciato_a = "il presente attestato viene conferito a"
    pdf.cell(200, 10, txt=testo_rilasciato_a, align = 'C', ln=True)
    pdf.ln(7)
    # blocco dati dinamici
    # dati anagrafici del candidato
    nome_cognome = nome + ' ' + cognome
    dati_utente = f"nato/a a {luogo_nascita} il {data_nascita} - Codice Fiscale: {codice_fiscale}"
    
    # Stampa nome e cognome
    pdf.set_font("Helvetica", 'B', size=32)
    pdf.set_text_color(235, 9, 16)
    pdf.set_margins(10, 10, 10)
    for riga in nome_cognome.splitlines():
        pdf.multi_cell(0, 10, riga, align = 'C', )
    pdf.ln(5)

    # stampa data_nascita,luogo,ecc...
    pdf.set_margins(10, 10, 10)
    pdf.set_font("Helvetica", "I", size=12)
    pdf.set_text_color(36, 35, 35)
    pdf.cell(200, 5, txt=dati_utente, align = 'C', ln=True)
    pdf.ln(5)  # Aggiunge una nuova linea
    
    # Stampa dettaglio azienda
    azienda = info['azienda'] + " - P.IVA: " + info['partita_iva']
    settore_ateco = info['settore_ateco']
    ente_formativo = info['ente_formativo']
    ente_certificatore = info['ente_certificatore']
    stampa_dati_azienda(pdf, azienda, settore_ateco, ente_formativo, ente_certificatore)

def disegna_linea(pdf):
    # DISEGNA LINEA ORIZZONTALE
    y_pos = pdf.get_y()
    # Impostazioni linea
    w = pdf.w
    line_length = (3/4) * w
    line_thickness = 0.5
    x_start = (w - line_length) / 2

    pdf.set_draw_color(36, 35, 35)
    pdf.set_draw_color(200, 200, 200)
    pdf.set_line_width(line_thickness)

    # Disegna la linea qualche mm sotto il testo (ad esempio 2 mm)
    pdf.line(x_start, y_pos + 2, x_start + line_length, y_pos + 2)
    pdf.ln(7)  # Spazio dopo header
    
def stampa_intestazione_pdf(pdf, image_path, header_text):
    # immagine ANPAL/UE in intestazione
    pdf.set_margins(10, 10, 10)
    pdf.ln(5)
    pdf.image(image_path, x=10, y=8, w=190)  # Modifica i valori di x, y, e w secondo necessità    
    pdf.ln(20)  # Spazio dopo l'immagine
    disegna_linea(pdf)
    pdf.ln(5)  # Spazio dopo header
    # INTESTAZIONE DEL DOCUMENTO - SUBITO SOTTO IMMAGINE HEADER
    pdf.set_text_color(235, 9, 16)
    pdf.set_font("Helvetica", "B", size=36)
    pdf.cell(193, 7, txt=header_text, align = 'C', ln=True)
    pdf.ln(5)  # Spazio dopo header
    pdf.set_text_color(36, 35, 35)
    disegna_linea(pdf)
    pdf.ln(3)  # Spazio dopo header
    
    

# Funzione che stampa "per la partecipazione al corso di formazione generale e specifica" ed il relativo corso passato come parametro
def stampa_denominazione_percorso_sviluppo(pdf, nome_corso_da_stampare):
    pdf.set_text_color(235, 9, 16)
    pdf.set_font("Arial", "B", size=24)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=nome_corso_da_stampare, align = 'C', ln=True)
    pdf.set_text_color(36, 35, 35)
    pdf.ln(6)

# Stampa testo in piu righe
def stampa_multicell(pdf, testo):
    pdf.add_font('DejaVu', '', './DejaVuSansCondensed.ttf', uni=True)
    pdf.set_font('DejaVu', '', 8)
    pdf.set_margins(10, 10, 10)
    for riga in testo.splitlines():
        pdf.multi_cell(0, 5, riga)
        # Stampa testo in piu righe
        
def stampa_multicell_grande(pdf, testo):
    pdf.set_font("Arial", "I", size=10)
    pdf.set_margins(20, 20, 20)
    for riga in testo.splitlines():
        pdf.multi_cell(0, 5, riga, align='C')
    pdf.set_margins(10, 10, 10)


# stampa paragrafo generica con header char size = 12
def stampa_paragrafo(pdf,header_str,text_string):
    #stampa header
    pdf.set_font("Helvetica", "BI", size=12)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=header_str, align = 'C', ln=True)
    pdf.ln(3)
    # stampa testo
    stampa_multicell(pdf, text_string)

# stampa paragrafo secondario con header char size = 10
def stampa_paragrafo_secondario(pdf, header_str, text_string):
    # stampa header
    pdf.set_font("Helvetica", "B", size=12)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=header_str, align = 'C', ln=True)
    pdf.ln(1)
    # stampa testo
    stampa_multicell(pdf, text_string)

# STAMPA LA TABELLA PER LE FIRME DI ACCETTAZIONE DEI VARI ENTI
def stampa_firme_accettazione(pdf, data_emissione):
    # Stampa Luogo e Data
    pdf.ln(2)
    disegna_linea(pdf)
    filigrana(pdf)
    pdf.set_margins(10, 10, 10)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(60, 10, 'Luogo e Data', 0, 0, 'C')
    pdf.set_font('Arial', 'I', 11)
    txt_date = 'Terni, lì ' + data_emissione
    pdf.text(25, pdf.get_y()+ 12, txt=txt_date)

    # Stampa Ente Formativo : 
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(60, 10, ' ', 0, 0, 'C')

    # Stampa Ente Certificatore
    pdf.cell(60, 10, 'Il Docente', 0, 1, 'C')
    
    # Ottieni la posizione y corrente per posizionare la firma sotto
    y_firma = pdf.get_y()  # aggiungi qualche mm di spazio
    larghezza_pagina = pdf.w
    margine_dx = pdf.r_margin
    # Calcola la posizione x della firma, allineata a destra come il testo
    nome_formatore = info['ente_formativo']
    img_w = 39  # larghezza immagine
    x_firma = larghezza_pagina - margine_dx - img_w - 15
    
    # Stampa la firma con coordinate dinamiche
    stampa_firma(pdf, nome_formatore, x_firma, y_firma)
    
    pdf.ln(15)
    pdf.cell(60, 10, ' ', 0, 0, 'C')
    
    pdf.cell(60, 10, ' ', 0, 0, 'C')
    # pdf.cell(60, 10, 'Il Direttore del Corso', 0, 1, 'C')

def stampa_firma(pdf, nome_formatore="", x=None, y=None):
    if not nome_formatore:
        return
    image_path = "firma_" + nome_formatore + ".png"
    img_w = 39
    img_h = 14
    # Se x o y non sono specificati, usa valori di default
    if x is None:
        x = (pdf.w - img_w) / 2
    if y is None:
        y = pdf.get_y()
    pdf.image(image_path, x=x, y=y, w=img_w, h=img_h)

# Funzione che salva il pdf e ci da in out i percorsi in cui sono stati salvati
def salvapdf(pdf, codice_fiscale, info, corso_id):
    # Salva il  pdf nella path /reports nominandolo come cf_siglacorso.pdf
    # Estrae la sigla del corso
    # Crea il percorso completo per il file PDF
    azienda = info['azienda']
    
    saved_file_name = f"{codice_fiscale}_{corso_id}.pdf"

    # Directory di test: output_path = os.path.join("prova_reports", azienda, saved_file_name)
    output_path = os.path.join("reports", azienda, saved_file_name)

    # Crea la cartella se non esiste
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    # Salva il PDF
    pdf.output(output_path, "F")
    print(f"Il file PDF è stato salvato in: {output_path}")


# funzione per convertire minuti in HH:MM
# serve per stampare il tempo dell'utente nel corso
def convert_minutes_to_hours_minutes(minutes):
    hours = int(minutes // 60)
    remaining_minutes = int(minutes % 60)
    return f"{hours:02d}:{remaining_minutes:02d}"

# funzione per convertire minuti in ore troncate
# serve per popolare la durata totale dei corsi
def convert_minutes_to_hours(minutes):
    hours = int(minutes // 60)
    return f"{hours}"

def stampa_riquadri(pdf):
    # DRAWING - Cornice + L agli angoli
    pdf.set_draw_color(200, 200, 200)
    w, h = pdf.w, pdf.h
    margin = 5
    # Rettangolo sottile interno
    thin_border = 0.5
    pdf.set_line_width(thin_border)
    pdf.rect(margin, margin, w - 2*margin, h - 2*margin)

    # Linee spesse "L"
    thick_line = 4
    length = 30
    pdf.set_draw_color(36, 35, 35)
    pdf.set_line_width(thick_line)

    # Angolo in basso a sinistra
    pdf.line(margin, h - margin, margin + length, h - margin)      # orizzontale
    pdf.line(margin, h - margin, margin, h - margin - length)      # verticale
    
    # L Interna - Config
    gap = 3   # margine interno tra le L
    thick_line = 3
    pdf.set_line_width(thick_line)
    pdf.set_draw_color(235, 9, 16)
    # Seconda L interna (ulteriore spostamento di gap)
    pdf.line(margin + 2*gap, h - margin - 2*gap, margin + length - 2*gap, h - margin - 2*gap)
    pdf.line(margin + 2*gap, h - margin - 2*gap, margin + 2*gap, h - margin - length + 2*gap)
    
    # Angolo in alto a destra
    pdf.set_draw_color(36, 35, 35)
    thick_line = 4
    pdf.set_line_width(thick_line)
    pdf.line(w - margin, margin, w - margin - length, margin)       # orizzontale verso sinistra
    pdf.line(w - margin, margin, w - margin, margin + length)       # verticale verso il basso  
    
    # L Interna - Config
    gap = 3   # margine interno tra le L
    thick_line = 3
    pdf.set_line_width(thick_line)
    pdf.set_draw_color(235, 9, 16)
    # coordinate spostate verso l'interno di 2*gap
    x_start = w - margin - 2*gap
    y_start = margin + 2*gap

    # lunghezza ridotta della L interna, ad esempio length - 3*gap
    reduced_length = length - 4*gap

    pdf.line(x_start, y_start, x_start - reduced_length, y_start)       # orizzontale più corta verso sinistra
    pdf.line(x_start, y_start, x_start, y_start + reduced_length)       # verticale più corta verso il basso
    # Colori e spessore
    pdf.set_draw_color(235, 9, 16)
    pdf.set_fill_color(235, 9, 16)
    pdf.set_line_width(1)

    margin = 3
    triangle_size = 20
    w, h = pdf.w, pdf.h

    # Triangolo in basso a destra
    points_br = [
        w - margin, h - margin,
        w - margin - triangle_size, h - margin,
        w - margin, h - margin - triangle_size
    ]
    polygon(pdf, points_br, style='F')

    # Triangolo in alto a sinistra
    points_tl = [
        margin, margin,
        margin + triangle_size, margin,
        margin, margin + triangle_size
    ]
    polygon(pdf, points_tl, style='F')
    
def polygon(pdf, points, style='D'):
    h = pdf.h
    k = pdf.k
    op = {'F': 'f', 'FD': 'b', 'DF': 'b'}.get(style, 's')
    points_str = ''
    for i in range(0, len(points), 2):
        x = points[i] * k
        y = (h - points[i+1]) * k
        if i == 0:
            points_str += f'{x:.2f} {y:.2f} m '
        else:
            points_str += f'{x:.2f} {y:.2f} l '
    pdf._out(points_str + op)

def filigrana(pdf):
    # dimensioni pagina
    page_w = pdf.w
    # dimensioni immagine da usare (puoi personalizzare)
    img_w = 30  # esempio: 100 mm di larghezza
    img_h = 30  # esempio: 100 mm di altezza
    # calcolo coordinate per centrare
    x = ((page_w - img_w) / 2)
    y = y = pdf.get_y()
    pdf.image("logo_square_autentica.png", x=x, y=y, w=img_w, h=img_h)
    
# Funzione per creare PDF per ogni codice fiscale
def crea_pdf(codice_fiscale, info):
    for corso in info['corsi']:
        # ----------------------------------------------------------------------------------------------------------------
        pdf = FPDF()
        pdf.add_page()
        
        # Stampa Header image + Attestato...
        image_path = './intestazione.png'
        testo_header_1 = "ATTESTATO di FORMAZIONE"
        stampa_intestazione_pdf(pdf, image_path, testo_header_1)
        
        # STAMPA DEL BLOCCO DATI UTENTE DA "Il presente Attestato viene conferito a:" FINO AI DATI ANAGRAFICI        
        blocco_dati_utente(pdf, info['data_nascita'], info['comune_nascita'], info['cognome'], info['nome'], codice_fiscale) # spaziatura verticale: pdf.ln(5)   
        pdf.ln()
        
        # STAMPA DETTAGLI CORSO
        id_corso = info['id_corso']
        csv_filename = "./svil_comp.csv"
        dettagli_percorso_formativo = csv_to_list(csv_filename, id_corso)
        
        if not dettagli_percorso_formativo or dettagli_percorso_formativo == None:
            print(f"Nessuna riga trovata con il titolo '{id_corso}'.")   

        # Stampa "ha frequentato / ha partecipato"
        if id_corso == 3:
            pdf.ln(4)
            pdf.set_font("Helvetica", "BI", size=14)
            testo_header = "ha frequentato il corso di formazione"
            pdf.cell(200, 7, txt=testo_header, align = 'C', ln=True)
            pdf.ln(8)
            
        else:
            pdf.ln(4)
            pdf.set_font("Helvetica", "BI", size=14)
            testo_header = "per la partecipazione al corso di formazione generale e specifica"
            pdf.cell(200, 7, txt=testo_header, align = 'C', ln=True)
            pdf.ln(8)
            
        # Stampa NOME DEL CORSO = value scritto su titolo_sessione_formativa(master) e su perco_titolo(svil_comp) trovato tramite il campo id_corso
        stampa_denominazione_percorso_sviluppo(pdf, dettagli_percorso_formativo[1]) # spaziatura verticale: pdf.ln(5)
        
        # Stampa "della durata di..."
        if id_corso == 3:
            header_str = f'della durata di {dettagli_percorso_formativo[4]} ore, in conformità al Reg. CE 852/2004'
            text_string = ""
            stampa_paragrafo(pdf, header_str, text_string)
            header_str = "Allegato II Cap. XII e s.m.i. ed ai sensi delle ulteriori normative nazionali e locali applicabili"
            text_string = ''.join((list(dettagli_percorso_formativo[2])))
            stampa_paragrafo(pdf, header_str, text_string)
            pdf.ln(2)
        
        else:
            # Stampa "ai sensi dell'art..."
            paragrafo_ai_sensi = "ai sensi dell'Art.37 del D. Lgs. n° 81/08 e s.m.i. e secondo l'Accordo Stato Regioni e Provincie Autonome di Trento e Bolzano del 21.12.2011"
            stampa_multicell_grande(pdf, paragrafo_ai_sensi)
            pdf.ln(2)
            # Stampa CONTENUTI DEL CORSO
                # STAMPA [4] = perco_durata
            header_str = f"della durata complessiva di {dettagli_percorso_formativo[4]} ore, tenutosi in modalità FAD, con i seguenti contenuti:"
                # STAMPA [2] = perco_obiettivo
            text_string = ''.join((list(dettagli_percorso_formativo[2]))) 
            stampa_paragrafo(pdf, header_str, text_string)
        
        pdf.ln(12)
        
        # Stampa diciture per firme accettazione
            # Stampiamo la data di emissione sincronizzata ad ora
        data_certificazione = datetime.now()
        data_certificazione = data_certificazione.strftime("%d-%m-%Y")
        stampa_firme_accettazione(pdf, data_certificazione)

        # Stampa Template Grafico
        stampa_riquadri(pdf)
        
        # SALVATAGGIO PDF
            # id_corso per distinzione nome files
        corso = dettagli_percorso_formativo[0]
        salvapdf(pdf, codice_fiscale, info, corso) 

# ----------------------------------------------------------------------------------------------------------------    
# ------------------------------------- FINE FUNZIONI ------------------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------
#########################################################################################################
        # Funzione che stampa i dettagli dell'utente relativi al percorso formativo: durata, ore totali e valutazione
        # passare in input il pdf, il dict del corso e la valutazione ottenuta al quiz 
        # momentaneamente la valutazione è fissa ma bisogna implementare la logica corretta    
        # valutazione = corso['valutazione_quiz']
        # DELETATA stampa_dettaglio_percorso_sviluppo(pdf, corso, valutazione)
        #########################################################################################################
        
        # IMPORTANTE: Dopo le modifiche del 20/05/2025, la chiave che lega master.xls a svil_comp.csv è:
        # il campo titolo_sessione_formativa di master è la chiave della colonna b di svil_comp.csv denominata perco_titolo
        # Quando dovremmo aggiungere una nuova tipologia di corso dovremmo aggiungere una riga al file svil_comp.csv
        # Nel file svil_comp bisogna compilare:
        # - la colonna perco_obiettivo con i dati relativi ai contenuti del corso
        # - la colonna perco_titolo con il titolo del corso
        # - la colonna perco_durata con il numero delle ore di durata del corso
        # Mentre nel file master.xlsx, quando si va ad aggiungere un utente che ha effettuato un corso bisogna
        # compilare come titolo_sessione_formativa lo stesso valore del campo perco_titolo del file svil_comp.csv
        # per assegnare all'utente un corso fatto e far stampare le denominazioni corrette sul PDF
        
        # DECODIFICA DELLA LISTA dettagli_percorso_formativo
        # [0] = perco_comp_id                   # [8] = comp_percorso
        # [1] = perco_titolo                    # [9] = comp_ref_areacompetenza
        # [2] = perco_obiettivo                 # [10] = comp_denominazione_competenza
        # [3] = perco_contenuto                 # [11] = comp_titolouc
        # [4] = perco_durata                    # [12] = comp_abilita
        # [5] = perco_tot_ore                   # [13] = comp_conoscenze
        # [6] = perco_destinatari
        # [7] = perco_mod_formativa
#
# ----------------------------------------------------------------------------------------------------------------  
# ------------------------------- INIZIO GENERAZIONE DEI FILE PDF ------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------
# Usa il percorso assoluto del file Excel
#  file_path = './dati.xlsx'

# file di test: file_path = './master_ridotta.xlsx' #file di prova tabella master ridotta

# file_path = './master_ridotta.xlsx' #file di prova tabella master ridotta
file_path = './master_ridotta.xlsx' # file definitivo

# Leggi il file Excel
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Crea un dizionario per memorizzare i dati
data = {}
n_stampe = 0
# Itera attraverso le righe del foglio Excel
for row in sheet.iter_rows(min_row=2, values_only=True):  # Salta l'intestazione
    moodle_institution, cognome, nome, codice_fiscale, data_di_nascita, comune_di_nascita, titolo_azione_con_id, titolo_sessione_formativa, Totaleorecorso, durata_min, stato_corso, valutazione_quiz, partita_iva, settore_ateco, ente_formativo, ente_certificatore, id_corso  = row

    if codice_fiscale is None:
        continue  # Salta la riga se il codice fiscale è None

    if codice_fiscale not in data:
        data[codice_fiscale] = {
            'azienda': moodle_institution,
            'nome': nome,
            'cognome': cognome,
            'data_nascita' : data_di_nascita.strftime("%d/%m/%Y"),
            'comune_nascita': comune_di_nascita,
            'partita_iva': partita_iva, 
            'settore_ateco': settore_ateco, 
            'ente_formativo': ente_formativo, 
            'ente_certificatore': ente_certificatore,
            'id_corso': id_corso,
            'corsi': []
        }
    
    data[codice_fiscale]['corsi'].append({
        'titolo_azione_con_id': titolo_azione_con_id,
        'titolo_sessione_formativa': titolo_sessione_formativa,
        'Totaleorecorso': Totaleorecorso,
        'durata_min': durata_min,
        'stato_corso': stato_corso,
        'valutazione_quiz': valutazione_quiz
    })
    n_stampe = n_stampe + 1

# Crea PDF per ogni codice fiscale
for codice_fiscale, info in data.items():
    crea_pdf(codice_fiscale, info)

print("N. totale stampe: "+str(n_stampe))