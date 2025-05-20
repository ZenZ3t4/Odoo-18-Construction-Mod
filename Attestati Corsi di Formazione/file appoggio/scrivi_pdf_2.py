from funzioni_reports import *
import openpyxl
from fpdf import FPDF
import os

# #####################################################

# Usa il percorso assoluto del file Excel
#  file_path = './dati.xlsx'
file_path = './peserico.xlsx' #file di Studio Peserico

# Leggi il file Excel
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Crea un dizionario per memorizzare i dati
data = {}

# Itera attraverso le righe del foglio Excel
for row in sheet.iter_rows(min_row=2, values_only=True):  # Salta l'intestazione
    moodle_institution, cognome, nome, codice_fiscale, data_di_nascita, comune_di_nascita, titolo_azione_con_id, titolo_sessione_formativa, Totaleorecorso, durata_min, stato_corso  = row

    if codice_fiscale is None:
        continue  # Salta la riga se il codice fiscale è None

    if codice_fiscale not in data:
        data[codice_fiscale] = {
            'azienda': moodle_institution,
            'nome': nome,
            'cognome': cognome,
            'data_nascita' : data_di_nascita,
            'comune_nascita': comune_di_nascita,
            'corsi': []
        }
    
    data[codice_fiscale]['corsi'].append({
        'titolo_azione_con_id': titolo_azione_con_id,
        'titolo_sessione_formativa': titolo_sessione_formativa,
        'Totaleorecorso': Totaleorecorso,
        'durata_min': durata_min,
        'stato_corso': stato_corso
    })

# Funzione per creare PDF per ogni codice fiscale
def crea_pdf(codice_fiscale, info):
    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # immagine ANPAL/UE in intestazione
    pdf.image('./intestazione.png', x=10, y=8, w=190)  # Modifica i valori di x, y, e w secondo necessità
    pdf.ln(20)  # Spazio dopo l'immagine
    
    # INTESTAZIONE DEL DOCUMENTO - SUBITO SOTTO IMMAGINE HEADER
    pdf.set_font("Arial", "B", size=12)
    testo_header_1 = "ALLEGATO 7 - ATTESTATO FINALE DI MESSA IN TRASPARENZA DELLE COMPETENZE"
    pdf.cell(193, 7, txt=testo_header_1, align = 'C', ln=True)
    # stampiamo in due righe altrimenti non ci entra in larghezza
    # pdf.set_font("Arial", "B", size=13)
    # testo_header_1 = " "
    # pdf.cell(200, 7, txt=testo_header_1, align = 'C', ln=True)
    pdf.ln(10)  # Spazio dopo header
    
    def stampa_dati_azienda(azienda, settore_ateco, ente_formativo, ente_certificatore):
        # BLOCCO DATI AZIENDA
        pdf.set_font("Arial", "B", size=16)
        azienda_header = "Dati identificativi dell'azienda"
        
        if len(settore_ateco) > 80:
            settore_ateco = settore_ateco[:80] + '...'
        
        # Dati dinamici azienda
        pdf.set_font("Helvetica", size=10)
        azienda_row_1 = f"AZIENDA: {azienda}"
        azienda_row_2 = f"Settore ateco: {settore_ateco}"
        azienda_row_3 = f"ENTE FORMATIVO: {ente_formativo}"
        azienda_row_4 = f"ENTE CERTIFICATORE: {ente_certificatore}"
        pdf.cell(200, 5, txt=azienda_header, ln=True)
        pdf.cell(200, 5, txt=azienda_row_1, ln=True)
        pdf.cell(200, 5, txt=azienda_row_2, ln=True)
        pdf.cell(200, 5, txt=azienda_row_3, ln=True)
        pdf.cell(200, 5, txt=azienda_row_4, ln=True)
        pdf.ln(10)  # Spazio dopo blocco dati azienda
        #--------------------------------------------------------------------------------
    
    azienda = info['azienda'] + " - P.IVA 09719910961"
    settore_ateco = '68.32.00 - AMMINISTRAZIONE DI CONDOMINI E GESTIONE DI BENI IMMOBILI PER CONTO TERZI'
    ente_formativo = "OBM SRL"
    ente_certificatore = "POLARIS HR"
    stampa_dati_azienda(azienda, settore_ateco, ente_formativo, ente_certificatore) # spaziatura: pdf.ln(10)

    # header metà attestato
    # header in grassetto
    pdf.set_font("Arial", "B", size=16)
    testo_header_2 = "ATTESTATO FINALE DI MESSA IN TRASPARENZA DELLE "
    pdf.cell(200, 7, txt=testo_header_2, align = 'C', ln=True)
    testo_header_2 = "COMPETENZE"
    pdf.cell(200, 7, txt=testo_header_2, align = 'C', ln=True)
    
    # header in corsivo
    pdf.set_font("Arial", "I", size=11)
    testo_subheader = "Relativa al percorso formativo realizzato nell'ambito di interventi finanziati dal Fondo Nuove Competenze"
    pdf.cell(193, 7, txt=testo_subheader, align = 'C', ln=True)
    pdf.ln(5)  # Aggiunge una nuova linea
    
    
    # STAMPA DEL BLOCCO DATI UTENTE DA "RILASCIATO A" FINO AI DATI ANAGRAFICI
    def blocco_dati_utente(data_nascita, luogo_nascita, cognome, nome, codice_fiscale):
        # BLOCCO DATI UTENTE
        # header in grassetto
        pdf.set_font("Arial", "B", size=16)
        testo_rilasciato_a = "RILASCIATO A"
        pdf.cell(200, 10, txt=testo_rilasciato_a, align = 'C', ln=True)

        # blocco dati dinamici
        dati_utente = f"{cognome} {nome} , nato a {luogo_nascita} il {data_nascita}  - C.F. {codice_fiscale}"
        # dati anagrafici del candidato
        pdf.set_font("Helvetica", size=10)
        pdf.cell(200, 5, txt=dati_utente, ln=True)
        pdf.ln(5)  # Aggiunge una nuova linea

    # in input passare in ordine data_nascita, luogo_nascita, info['cognome'], info['nome'], codice_fiscale
    data_nascita = info['data_nascita']
    comune_nascita = info['comune_nascita']
    blocco_dati_utente(data_nascita, comune_nascita, info['cognome'], info['nome'], codice_fiscale) # spaziatura verticale: pdf.ln(5)
    
    # Header Corsi effettuati
    pdf.cell(200, 10, txt="Corsi Effettuati", ln=True)
    pdf.ln()

    # funzione per convertire minuti in ore
    def convert_minutes_to_hours_minutes(minutes):
        hours = int(minutes // 60)
        remaining_minutes = int(minutes % 60)
        return f"{hours:02d}:{remaining_minutes:02d}"
    
    def convert_minutes_to_hours(minutes):
        hours = int(minutes // 60)
        remaining_minutes = int(minutes % 60)
        return f"{hours:02d}"
    
    def stampa_denominazione_percorso_sviluppo(denominazione_corso):
            pdf.set_font("Arial", "B", size=13)
            testo_header = "DENOMINAZIONE DEL PERCORSO DI SVILUPPO DELLE COMPETENZE:"
            pdf.cell(200, 7, txt=testo_header, align = 'C', ln=True)
            pdf.cell(200, 7, txt=denominazione_corso, align = 'C', ln=True)
            pdf.ln(5)  # Aggiunge una nuova linea

    # inizio stampa dettaglio corsi
    for corso in info['corsi']:
        # creazione blocco header
        # Header denominazione corso
        denominazione_corso = corso['titolo_sessione_formativa']
        stampa_denominazione_percorso_sviluppo(denominazione_corso) # spaziatura verticale: pdf.ln(5)
        
        # creazione stringhe
        valutazione = "14/15"
        esito_corso = corso['stato_corso']
        durata_ore = convert_minutes_to_hours_minutes(corso['durata_min'])
        totale_ore_corso = convert_minutes_to_hours(corso['Totaleorecorso'])
        testo_intro_corsi = f"Percorso: {corso['titolo_azione_con_id']}"
        testo_corsi_2 = f"- Durata del percorso di {totale_ore_corso} ore di cui {totale_ore_corso} ore di FAD asincrona. Ore effettive svolte: {durata_ore} ore."
        testo_corsi_3 = f"- Corso {esito_corso}  -  Valutazione finale: {valutazione}  -  valutazione effettuata tramite quiz."

        #stampa stringhe
        pdf.set_font("Helvetica", size=10)
        pdf.cell(200, 5, txt=testo_intro_corsi, ln=True)
        pdf.cell(200, 5, txt=testo_corsi_2, ln=True)
        pdf.cell(200, 5, txt=testo_corsi_3, ln=True)
        pdf.ln(5)
        
        
        
    # d_svil_comp è un dict che contiene i dettagli delle varie competenze acquisite nei corsi
    # d_svil_comp = csv_to_dict("svil_comp.csv")
    #print(d_svil_comp.keys())
    sigla_corso = list(denominazione_corso)
    sigla_corso = str(sigla_corso[6])
    # Salva il PDF
    pdf.output(f"./reports/{codice_fiscale}.pdf", "F")
    output_path = os.path.abspath(path="reports") + f"\\{codice_fiscale}_{sigla_corso}.pdf"
    print(f"Il file PDF è stato salvato in: {output_path}")

# Crea PDF per ogni codice fiscale
for codice_fiscale, info in data.items():
    crea_pdf(codice_fiscale, info)
