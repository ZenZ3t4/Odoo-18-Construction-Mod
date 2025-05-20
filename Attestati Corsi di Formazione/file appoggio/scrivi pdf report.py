import fpdf
import openpyxl
import os

# Usa il percorso assoluto del file Excel
file_path = './dati.xlsx'
workbook = openpyxl.load_workbook(file_path)

# Seleziona il foglio attivo
sheet = workbook.active

# Leggi i dati dal file Excel e crea delle liste conteneti i vari dati estratti
dati_anagrafici = {f'{(sheet.cell(1, 1))}': []}
                   
corsi = []
voti = []
superamento = []
esito = []

for row in range(2, sheet.max_row + 1):
    corsi.append(sheet.cell(row=row, column=4).value)
    voti.append(sheet.cell(row=row, column=5).value)
    superamento.append(sheet.cell(row=row, column=6).value)
    
    # generazione di una colonna con parametri condizionali dallo script
    testo = "non superato"
    if voti[-1] in ["12/15", "13/15", "14/15"]:
        testo = "ok"
    esito.append(testo)
    
# Funzione per stampare una riga della tabella
def stampa_riga(colonne):
    for colonna in colonne:
        pdf.cell(50, 10, txt=str(colonna), border=1, align='C')
    pdf.ln()

# +*************** Creazione del PDF in memoria **************+
# Creare un oggetto PDF
pdf = fpdf.FPDF()

# Aggiungere una pagina
pdf.add_page()

# Impostare il font
pdf.set_font("Arial", size=12)

# BOX INTESTAZIONE DEI DATI ANAGRAFICI DELL'AZIENDA [moodle_institution]
pdf.cell(50, 10, txt=str(colonna), border=1, align='C')

# +*************** Procedura di stampa nella tabella **************+
# Stampare l'intestazione della tabella
colnames = ["Corso", "Votazione", "Superato", "Prova"]
stampa_riga(colnames)

# Stampare le righe della tabella
for i in range(len(corsi)):
    riga = [corsi[i], voti[i], superamento[i], esito[i]]
    stampa_riga(riga)
# +*************** ********************************* **************+

# Salvare il PDF
output_path = os.path.join(os.getcwd(), "prova.pdf")
pdf.output(output_path, "F")

print(f"Il file PDF è stato salvato in: {output_path}")
