from funzioni_reports import *
import openpyxl
from fpdf import FPDF
from datetime import datetime
import os
import logging

log_dir = './logs'
os.makedirs(log_dir, exist_ok=True)
log_file_path = os.path.join(log_dir, 'app.log')
logging.basicConfig(
    filename=log_file_path,
    filemode='a',
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level=logging.INFO
)

# #####################################################
# -----------------------------------------
# FUNZIONI UTILIZZATE NEL CODICE
# -----------------------------------------
def stampa_dati_azienda(pdf, dettagli_azienda):
    # Dati dinamici azienda
    pdf.set_font("Helvetica", "I", size=11)
    #azienda_row_1 = f"{azienda} - Settore ateco: {settore_ateco}"
    azienda_row_1 = f"collaboratore dell'azienda {dettagli_azienda['ragione_sociale']} - P.IVA: {dettagli_azienda['partita_iva']}"
    pdf.cell(200, 5, txt=azienda_row_1, align = 'C', ln=True)
    #--------------------------------------------------------------------------------
    
def blocco_dati_utente(pdf, data_nascita, luogo_nascita, nominativo, codice_fiscale):
    # BLOCCO DATI UTENTE
    # header in grassetto
    pdf.set_font("Helvetica", "BI", size=14)
    pdf.set_margins(10, 10, 10)
    testo_rilasciato_a = "il presente attestato viene conferito a"
    pdf.cell(200, 10, txt=testo_rilasciato_a, align = 'C', ln=True)
    pdf.ln(7)
    # blocco dati dinamici
    # dati anagrafici del candidato
    dati_utente = f"nato/a a {luogo_nascita} il {data_nascita} - Codice Fiscale: {codice_fiscale}"
    
    # Stampa nome e cognome
    pdf.set_font("Helvetica", 'B', size=32)
    pdf.set_text_color(235, 9, 16)
    pdf.set_margins(10, 10, 10)
    for riga in nominativo.splitlines():
        pdf.multi_cell(0, 10, riga, align = 'C', )
    pdf.ln(5)

    # stampa data_nascita,luogo,ecc...
    pdf.set_margins(10, 10, 10)
    pdf.set_font("Helvetica", "I", size=12)
    pdf.set_text_color(36, 35, 35)
    pdf.cell(200, 5, txt=dati_utente, align = 'C', ln=True)
    pdf.ln(5)  # Aggiunge una nuova linea

def disegna_linea(pdf):
    # DISEGNA LINEA ORIZZONTALE
    y_pos = pdf.get_y()
    # Impostazioni linea
    w = pdf.w
    line_length = (3/4) * w
    line_thickness = 0.5
    x_start = (w - line_length) / 2

    pdf.set_draw_color(36, 35, 35)
    pdf.set_draw_color(200, 200, 200)
    pdf.set_line_width(line_thickness)

    # Disegna la linea qualche mm sotto il testo (ad esempio 2 mm)
    pdf.line(x_start, y_pos + 2, x_start + line_length, y_pos + 2)
    pdf.ln(7)  # Spazio dopo header
    
def stampa_intestazione_pdf(pdf, image_path, header_text):
    # immagine ANPAL/UE in intestazione
    pdf.set_margins(10, 10, 10)
    pdf.ln(5)
    pdf.image(image_path, x=10, y=8, w=190)  # Modifica i valori di x, y, e w secondo necessità    
    pdf.ln(20)  # Spazio dopo l'immagine
    disegna_linea(pdf)
    pdf.ln(5)  # Spazio dopo header
    # INTESTAZIONE DEL DOCUMENTO - SUBITO SOTTO IMMAGINE HEADER
    pdf.set_text_color(235, 9, 16)
    pdf.set_font("Helvetica", "B", size=36)
    pdf.cell(193, 7, txt=header_text, align = 'C', ln=True)
    pdf.ln(5)  # Spazio dopo header
    pdf.set_text_color(36, 35, 35)
    disegna_linea(pdf)
    pdf.ln(3)  # Spazio dopo header

# Funzione che stampa "per la partecipazione al corso di formazione generale e specifica" ed il relativo corso passato come parametro
def stampa_denominazione_percorso_sviluppo(pdf, nome_corso_da_stampare):
    pdf.set_text_color(235, 9, 16)
    pdf.set_font("Arial", "B", size=24)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=nome_corso_da_stampare, align = 'C', ln=True)
    pdf.set_text_color(36, 35, 35)
    pdf.ln(6)

# Stampa testo in piu righe
def stampa_multicell(pdf, testo):
    pdf.add_font('DejaVu', '', './DejaVuSansCondensed.ttf', uni=True)
    pdf.set_font('DejaVu', '', 8)
    pdf.set_margins(10, 10, 10)
    for riga in testo.splitlines():
        pdf.multi_cell(0, 5, riga)
        # Stampa testo in piu righe
        
def stampa_multicell_grande(pdf, testo):
    pdf.set_font("Arial", "I", size=10)
    pdf.set_margins(20, 20, 20)
    for riga in testo.splitlines():
        pdf.multi_cell(0, 5, riga, align='C')
    pdf.set_margins(10, 10, 10)


# stampa paragrafo generica con header char size = 12
def stampa_paragrafo(pdf,header_str,text_string):
    #stampa header
    pdf.set_font("Helvetica", "BI", size=12)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=header_str, align = 'C', ln=True)
    pdf.ln(3)
    # stampa testo
    stampa_multicell(pdf, text_string)

# stampa paragrafo secondario con header char size = 10
def stampa_paragrafo_secondario(pdf, header_str, text_string):
    # stampa header
    pdf.set_font("Helvetica", "B", size=12)
    pdf.set_margins(10, 10, 10)
    pdf.cell(200, 7, txt=header_str, align = 'C', ln=True)
    pdf.ln(1)
    # stampa testo
    stampa_multicell(pdf, text_string)

# STAMPA LA TABELLA PER LE FIRME DI ACCETTAZIONE DEI VARI ENTI
def stampa_firme_accettazione(pdf, data_emissione, docente):
    # Stampa Luogo e Data
    pdf.ln(2)
    disegna_linea(pdf)
    stampa_timbro(pdf)
    pdf.set_margins(10, 10, 10)
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(60, 10, 'Luogo e Data', 0, 0, 'C')
    pdf.set_font('Arial', 'I', 11)
    txt_date = 'Terni, lì ' + data_emissione
    pdf.text(25, pdf.get_y()+ 12, txt=txt_date)

    # Stampa Ente Formativo : 
    pdf.set_font('Arial', 'B', 12)
    pdf.cell(60, 10, ' ', 0, 0, 'C')

    # Stampa Ente Certificatore
    pdf.cell(60, 10, 'Il Docente', 0, 1, 'C')
    
    # Ottieni la posizione y corrente per posizionare la firma sotto
    y_firma = pdf.get_y()  # aggiungi qualche mm di spazio
    larghezza_pagina = pdf.w
    margine_dx = pdf.r_margin
    # Calcola la posizione x della firma, allineata a destra come il testo
    nome_formatore = docente
    img_w = 39  # larghezza immagine
    x_firma = larghezza_pagina - margine_dx - img_w - 15
    
    # Stampa la firma con coordinate dinamiche
    stampa_firma(pdf, nome_formatore, x_firma, y_firma)
    
    pdf.ln(15)
    pdf.cell(60, 10, ' ', 0, 0, 'C')
    
    pdf.cell(60, 10, ' ', 0, 0, 'C')
    # pdf.cell(60, 10, 'Il Direttore del Corso', 0, 1, 'C')

def stampa_firma(pdf, nome_formatore="", x=None, y=None):
    if not nome_formatore:
        return
    image_path = "firma_" + nome_formatore + ".png"
    img_w = 39
    img_h = 14
    # Se x o y non sono specificati, usa valori di default
    if x is None:
        x = (pdf.w - img_w) / 2
    if y is None:
        y = pdf.get_y()
    pdf.image(image_path, x=x, y=y, w=img_w, h=img_h)

# Funzione che salva il pdf e ci da in out i percorsi in cui sono stati salvati
def salvapdf(pdf, codice_fiscale, info, corso_id, ragione_sociale):
    # Salva il  pdf nella path /reports nominandolo come cf_siglacorso.pdf
    # Estrae la sigla del corso
    # Crea il percorso completo per il file PDF
    
    saved_file_name = f"{codice_fiscale}_{corso_id}.pdf"

    # Directory di test: output_path = os.path.join("prova_reports", azienda, saved_file_name)
    output_path = os.path.join("new_reports", ragione_sociale, saved_file_name)

    # Crea la cartella se non esiste
    output_dir = os.path.dirname(output_path)
    os.makedirs(output_dir, exist_ok=True)
    # Salva il PDF
    pdf.output(output_path, "F")
    print(f"Il file PDF è stato salvato in: {output_path}")
    logging.info(f"Il file PDF è stato salvato in: {output_path}")

# Composizione template grafico del PDF

# Cornice e riquadro + triangoli
def stampa_riquadri(pdf):
    # DRAWING - Cornice + L agli angoli
    pdf.set_draw_color(200, 200, 200)
    w, h = pdf.w, pdf.h
    margin = 5
    # Rettangolo sottile interno
    thin_border = 0.5
    pdf.set_line_width(thin_border)
    pdf.rect(margin, margin, w - 2*margin, h - 2*margin)

    # Linee spesse "L"
    thick_line = 4
    length = 30
    pdf.set_draw_color(36, 35, 35)
    pdf.set_line_width(thick_line)

    # Angolo in basso a sinistra
    pdf.line(margin, h - margin, margin + length, h - margin)      # orizzontale
    pdf.line(margin, h - margin, margin, h - margin - length)      # verticale
    
    # L Interna - Config
    gap = 3   # margine interno tra le L
    thick_line = 3
    pdf.set_line_width(thick_line)
    pdf.set_draw_color(235, 9, 16)
    # Seconda L interna (ulteriore spostamento di gap)
    pdf.line(margin + 2*gap, h - margin - 2*gap, margin + length - 2*gap, h - margin - 2*gap)
    pdf.line(margin + 2*gap, h - margin - 2*gap, margin + 2*gap, h - margin - length + 2*gap)
    
    # Angolo in alto a destra
    pdf.set_draw_color(36, 35, 35)
    thick_line = 4
    pdf.set_line_width(thick_line)
    pdf.line(w - margin, margin, w - margin - length, margin)       # orizzontale verso sinistra
    pdf.line(w - margin, margin, w - margin, margin + length)       # verticale verso il basso  
    
    # L Interna - Config
    gap = 3   # margine interno tra le L
    thick_line = 3
    pdf.set_line_width(thick_line)
    pdf.set_draw_color(235, 9, 16)
    # coordinate spostate verso l'interno di 2*gap
    x_start = w - margin - 2*gap
    y_start = margin + 2*gap

    # lunghezza ridotta della L interna, ad esempio length - 3*gap
    reduced_length = length - 4*gap

    pdf.line(x_start, y_start, x_start - reduced_length, y_start)       # orizzontale più corta verso sinistra
    pdf.line(x_start, y_start, x_start, y_start + reduced_length)       # verticale più corta verso il basso
    # Colori e spessore
    pdf.set_draw_color(235, 9, 16)
    pdf.set_fill_color(235, 9, 16)
    pdf.set_line_width(1)

    margin = 3
    triangle_size = 20
    w, h = pdf.w, pdf.h

    # Triangolo in basso a destra
    points_br = [
        w - margin, h - margin,
        w - margin - triangle_size, h - margin,
        w - margin, h - margin - triangle_size
    ]
    polygon(pdf, points_br, style='F')

    # Triangolo in alto a sinistra
    points_tl = [
        margin, margin,
        margin + triangle_size, margin,
        margin, margin + triangle_size
    ]
    polygon(pdf, points_tl, style='F')

# Funzione per creare graficamente i triangoli
def polygon(pdf, points, style='D'):
    h = pdf.h
    k = pdf.k
    op = {'F': 'f', 'FD': 'b', 'DF': 'b'}.get(style, 's')
    points_str = ''
    for i in range(0, len(points), 2):
        x = points[i] * k
        y = (h - points[i+1]) * k
        if i == 0:
            points_str += f'{x:.2f} {y:.2f} m '
        else:
            points_str += f'{x:.2f} {y:.2f} l '
    pdf._out(points_str + op)

# Funzione per stampare il logo Authentica s.p.a. al center-bottom nel PDF
def stampa_timbro(pdf):
    # dimensioni pagina
    page_w = pdf.w
    # dimensioni immagine da usare (puoi personalizzare)
    img_w = 30  # esempio: 100 mm di larghezza
    img_h = 30  # esempio: 100 mm di altezza
    # calcolo coordinate per centrare
    x = ((page_w - img_w) / 2)
    y = y = pdf.get_y()
    pdf.image("logo_square_autentica.png", x=x, y=y, w=img_w, h=img_h)

# Estrae ragione_sociale, p_iva, ateco da aziende.xlsx 
#   Key: id_azienda | colonna A | es: authentica, sarca, enerstreet, ecc...
#   Restituisce una tabella dettagli_azienda con i vari dati estratti
def estrai_dettagli_azienda(nome_file, id_azienda):
    # Apri il workbook
    workbook = openpyxl.load_workbook(nome_file)
    sheet = workbook.active  # Ottieni il foglio attivo
    # Trova gli header delle colonne in prima riga
    header = [cell.value for cell in sheet[1]]
    try:
        idx_azienda_id = header.index('azienda_id') + 1
        idx_ragione_sociale = header.index('ragione_sociale') + 1
        idx_partita_iva = header.index('partita_iva') + 1
        idx_codice_ateco = header.index('codice_ateco') + 1
    except ValueError:
        print("Una o più colonne necessarie non sono state trovate nel file Excel.")
        logging.error("Una o più colonne necessarie non sono state trovate nel file Excel.")
        return None
    
    dettagli_azienda = {}

    # Itera sulle righe del foglio (partendo dalla seconda riga per saltare l'intestazione)
    for row in sheet.iter_rows(min_row=2):
        if row[idx_azienda_id - 1].value == id_azienda:
            dettagli_azienda['ragione_sociale'] = row[idx_ragione_sociale - 1].value
            dettagli_azienda['partita_iva'] = row[idx_partita_iva - 1].value
            dettagli_azienda['codice_ateco'] = row[idx_codice_ateco - 1].value
            return dettagli_azienda  # Restituisci i dettagli non appena trovi l'azienda

    return None  # Restituisci None se l'azienda non viene trovata

# Funzione per creare PDF per ogni codice fiscale
def crea_pdf(codice_fiscale, info):
    for corso in info['corsi']:
        # ----------------------------------------------------------------------------------------------------------------
        pdf = FPDF()
        pdf.add_page()
        
        # Stampa Header image + Attestato...
        image_path = './intestazione.png'
        testo_header_1 = "ATTESTATO di FORMAZIONE"
        stampa_intestazione_pdf(pdf, image_path, testo_header_1)
        
        # STAMPA DEL BLOCCO DATI UTENTE DA "Il presente Attestato viene conferito a:" FINO AI DATI ANAGRAFICI        
        blocco_dati_utente(pdf, info['data_nascita'], info['comune'], info['nominativo'], codice_fiscale) # spaziatura verticale: pdf.ln(5)   
        # Stampa dettaglio azienda
        # Esempio di utilizzo:
        id_azienda = info['azienda']
        filename = './report_gen_data/db/aziende.xlsx'
        dettagli_azienda = estrai_dettagli_azienda(filename, id_azienda)
        if dettagli_azienda is None:
            print(f"Az:{id_azienda} ; CF: {codice_fiscale}; Nessuna azienda trovata con ID: {id_azienda};")
            logging.error(f"Az:{id_azienda} ; CF: {codice_fiscale}; Nessuna azienda trovata con ID: {id_azienda};")
            dettagli_azienda = {
                'ragione_sociale': 'Dati aziendali non disponibili',
                'partita_iva': 'N/A',
                'codice_ateco': 'N/A'
            }
            
        stampa_dati_azienda(pdf, dettagli_azienda)
        pdf.ln()
        
        # STAMPA DETTAGLI CORSO
        id_corso = corso['id_corso']
        if not id_corso or id_corso == None:
            print(f"Az:{id_azienda} ; CF: {codice_fiscale}; id_corso assegnato non trovato '{id_corso}'.")
            logging.error(f"Az:{id_azienda} ; CF: {codice_fiscale}; id_corso assegnato non trovato '{id_corso}'.")
        
        csv_filename = "./svil_comp.csv"
        dettagli_percorso_formativo = csv_to_list(csv_filename, id_corso)
        
        if not dettagli_percorso_formativo or dettagli_percorso_formativo == None:
            print(f"Az:{id_azienda} ; CF: {codice_fiscale}; Nessuna riga trovata con il titolo '{id_corso}'.")
            logging.error(f"Az:{id_azienda} ; CF: {codice_fiscale}; Nessuna riga trovata con il titolo '{id_corso}'.")
            return

        # Stampa "ha frequentato / ha partecipato"
        if id_corso == 3:
            pdf.ln(4)
            pdf.set_font("Helvetica", "BI", size=14)
            testo_header = "ha frequentato il corso di formazione"
            pdf.cell(200, 7, txt=testo_header, align = 'C', ln=True)
            pdf.ln(8)
            
        else:
            pdf.ln(4)
            pdf.set_font("Helvetica", "BI", size=14)
            testo_header = "per la partecipazione al corso di formazione generale e specifica"
            pdf.cell(200, 7, txt=testo_header, align = 'C', ln=True)
            pdf.ln(8)
            
        # Stampa NOME DEL CORSO = value scritto su titolo_sessione_formativa(master) e su perco_titolo(svil_comp) trovato tramite il campo id_corso
        stampa_denominazione_percorso_sviluppo(pdf, dettagli_percorso_formativo[1]) # spaziatura verticale: pdf.ln(5)
        
        # Stampa "della durata di..."
        if id_corso == 3:
            header_str = f'della durata di {dettagli_percorso_formativo[4]} ore, in conformità al Reg. CE 852/2004'
            text_string = ""
            stampa_paragrafo(pdf, header_str, text_string)
            header_str = "Allegato II Cap. XII e s.m.i. ed ai sensi delle ulteriori normative nazionali e locali applicabili"
            text_string = ''.join((list(dettagli_percorso_formativo[2])))
            stampa_paragrafo(pdf, header_str, text_string)
            pdf.ln(2)
        
        else:
            # Stampa "ai sensi dell'art..."
            paragrafo_ai_sensi = "ai sensi dell'Art.37 del D. Lgs. n° 81/08 e s.m.i. e secondo l'Accordo Stato Regioni e Provincie Autonome di Trento e Bolzano del 21.12.2011"
            stampa_multicell_grande(pdf, paragrafo_ai_sensi)
            pdf.ln(2)
            # Stampa CONTENUTI DEL CORSO
                # STAMPA [4] = perco_durata
            header_str = f"della durata complessiva di {dettagli_percorso_formativo[4]} ore, tenutosi in modalità FAD, con i seguenti contenuti:"
                # STAMPA [2] = perco_obiettivo
            text_string = ''.join((list(dettagli_percorso_formativo[2]))) 
            stampa_paragrafo(pdf, header_str, text_string)
        
        pdf.ln(12)
        
        # Stampa diciture per firme accettazione
            # Stampiamo la data di emissione sincronizzata ad ora
        data_certificazione = datetime.now()
        data_certificazione = data_certificazione.strftime("%d-%m-%Y")
        stampa_firme_accettazione(pdf, data_certificazione, corso['docente'])

        # Stampa Template Grafico
        stampa_riquadri(pdf)
        
        # SALVATAGGIO PDF
            # id_corso per distinzione nome files
        corso = dettagli_percorso_formativo[0]
        ragione_sociale = dettagli_azienda['ragione_sociale']
        salvapdf(pdf, codice_fiscale, info, corso, ragione_sociale) 
        
# ----------------------------------------------------------------------------------------------------------------    
# -------------------------------- FINE FUNZIONE GENERAZIONE PDF -------------------------------------------------
# ----------------------------------------------------------------------------------------------------------------
# ------------------ FUNZIONI ESTRAZIONE DATI ANAGRAFICI DA TABELLE AUTHENTICA -----------------------------------  
# ----------------------------------------------------------------------------------------------------------------
# Carica i codici catastali da file Excel in un dizionario.
# Returns:
#       dict: Dizionario {codice_catastale: (comune, provincia)}
def carica_codici_catastali(percorso_file_excel):
    try:
        wb = openpyxl.load_workbook(percorso_file_excel)
        sheet = wb.active

        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(sheet[1])}
        idx_catastale = headers.get("Codice Catastale del comune")
        idx_comune = headers.get("Denominazione in italiano")
        idx_provincia = headers.get("Sigla automobilistica")

        if idx_catastale is None or idx_comune is None or idx_provincia is None:
            raise ValueError("Colonne richieste non trovate")

        mappa_catastali = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            codice = str(row[idx_catastale]).strip().upper()
            comune = str(row[idx_comune]).strip()
            provincia = str(row[idx_provincia]).strip()
            mappa_catastali[codice] = (comune, provincia)

        return mappa_catastali

    except Exception as e:
        print("Errore nella lettura del file:", e)
        return {}

# Decodifica comune e provincia da un dizionario già caricato.  
#    Args:
#        codice_fiscale (str): Il codice fiscale (16 caratteri).
#        dizionario_catastali (dict): Dizionario con chiave codice catastale.
    
#    Returns:
#        tuple: (comune, provincia) oppure (None, None)
def decodifica_da_dizionario(cf, mappa_catastali):
    if len(cf) != 16:
        return None, None

    codice_catastale = cf[11:15].upper()
    comune, provincia = mappa_catastali.get(codice_catastale, (None, None))

    # Se provincia è vuota, e il codice inizia per Z, consideralo estero
    if codice_catastale.startswith("Z") and provincia is None:
        provincia = "EE"  # Estero

    return comune, provincia

    
# Funzione che estrae data di nascita da codice fiscale e restituisce gg/mm/aaaa
def estrai_data_nascita_da_codice_fiscale(codice_fiscale):
    mesi_codice = 'ABCDEHLMPRST'  # Mesi: A=Gen, B=Feb, ..., T=Dic
    anno = int(codice_fiscale[6:8])
    mese = mesi_codice.index(codice_fiscale[8]) + 1
    giorno = int(codice_fiscale[9:11])
    # Correzione giorno per donne
    if giorno > 40:
        giorno -= 40
    # Determina il secolo
    if anno < 40:
        anno += 2000
    else:
        anno += 1900
    data = f"{giorno}/{mese}/{anno}"
    return data

def crea_nominativo(record):
    surname = record.get('cognome', '')
    name = record.get('nome', '')  # Se disponibile, altrimenti metti ''
    return f"{surname} {name}".strip()

def trova_file_per_azienda(folder, azienda):
    for filename in os.listdir(folder):
        if (filename.endswith(".xls") or filename.endswith(".xlsx")) and azienda in filename:
            return os.path.join(folder, filename)
    return None

# ----------------------------------------------------------------------------------------------------------------
# Function - FOR generazione PDF
# ----------------------------------------------------------------------------------------------------------------
def genera_PDF(data, azienda_da_gen):
    n_errori = 0
    n_stampe = 0
    n_non_generati = 0
    for codice_fiscale, info in data.items():
        # Supponiamo che 'corsi' sia una lista di dizionari con 'id_corso'
        # Se vuoi generare un PDF per ogni corso presente:
        for corso in info.get('corsi', []):
            id_corso = corso.get('id_corso')
            if not id_corso:  # id_corso è None, stringa vuota o valore falsy
                n_non_generati += 1
                logging.warning(f"ID corso {id_corso} mancante per azienda '{azienda_da_gen}', codice fiscale '{codice_fiscale}'. PDF non generato.")
                print(f"Attenzione: ID corso mancante per azienda '{azienda_da_gen}', codice fiscale '{codice_fiscale}'. PDF non generato.")
                continue
            try:
                # Se id_corso è presente, genera il PDF
                crea_pdf(codice_fiscale, info)
                n_stampe += 1
            except Exception as e:
                n_errori += 1
                print(f"Errore nel creare PDF per {codice_fiscale}: {e}")
                logging.error(f"Errore nel creare PDF per {codice_fiscale}: {e}")

    print("N. totale stampe: "+str(n_stampe))
    print(f"N. PDF non generati per mancanza id_corso: {n_non_generati}")
    print("N. ERRORI CRITICI [except for codice_fiscale, info in data.items]: "+str(n_errori))
    logging.info("N. totale stampe: "+str(n_stampe))
    logging.info(f"N. PDF non generati per mancanza id_corso: {n_non_generati}")
    logging.info("N. ERRORI CRITICI [except for codice_fiscale, info in data.items]: "+str(n_errori))
# ------------------------------------------------------------------------------------------
# INIZIO GENERAZIONE PDF

# CARTELLA FILE MASTER,AZIENDA E aziende.xlsx: folder_path | cartella attuale: ./report_gen_data
# WORKBOOK MASTER: file_master_path
# WORKBOOK AZIENDA: filname che contiene azienda_da_gen stringa come da tabella aziende.xlsx
#                   ad esempio azienda_da_gen = "sarca":
#                           - aprirà il file che nel nome contiene la stringa "sarca"
#                           - nel file aziende.xlsx verranno estratti i dati di id_azienda = "sarca" [Colonna A del file]
# Variabili iniziali
logging.info("------------------------------------------------------------------------------------------- ")
logging.info("-------------------------------STARTING SESSION LOG --------------------------------------- ")
folder_path = "./report_gen_data"
file_master_path = './report_gen_data/db/new_master.xlsx'
file_codici_catastali = "./report_gen_data/db/codici_catastali_comuni.xlsx"
azienda_da_gen = input("Inserisci id azienda da aziende.xlsx: ")

# 1. Apri workbook master
workbook = openpyxl.load_workbook(file_master_path)
sheet_master = workbook.active

# 2. Trova file azienda corrispondente
file_azienda_path = trova_file_per_azienda(folder_path, azienda_da_gen)

if file_azienda_path is None:
    raise FileNotFoundError(f"Nessun file Excel trovato in '{folder_path}' contenente '{azienda_da_gen}' nel nome.")
print(f"File trovato: {file_azienda_path}")
logging.info(f"File trovato: {file_azienda_path}")

# 3. Carica workbook azienda
wb_azienda = openpyxl.load_workbook(file_azienda_path)
sheet_azienda = wb_azienda.active

# 4. Leggi intestazioni master e crea dizionario codici fiscali esistenti
headers_master = [cell.value for cell in next(sheet_master.iter_rows(min_row=1, max_row=1))]
idx_cf_master = headers_master.index('codice_fiscale')

# Dizionario per memorizzare i dati master, come nel tuo codice
data = {}
n_stampe = 0
n_errori = 0
dict_codici_catastali = carica_codici_catastali(file_codici_catastali)

for row in sheet_master.iter_rows(min_row=2, values_only=True):
    cf = row[idx_cf_master]
    if cf is None:
        continue

    nominativo = row[headers_master.index('nominativo')]
    comune = row[headers_master.index('comune')]
    provincia = row[headers_master.index('provincia')]
    sesso = row[headers_master.index('sesso')]
    e_mail = row[headers_master.index('e_mail')]
    id_azienda = row[headers_master.index('id_azienda')]
    id_corso = row[headers_master.index('id_corso')]
    docente = row[headers_master.index('docente')]
    
    if cf not in data:
        data[cf] = {
            'nominativo': nominativo,
            'data_nascita': estrai_data_nascita_da_codice_fiscale(cf),
            'comune': comune,
            'provincia': provincia,
            'sesso': sesso,
            'e_mail': e_mail,
            'azienda': id_azienda,
            'corsi': []
        }
    data[cf]['corsi'].append({
        'id_corso': id_corso,
        'docente': docente   # associato al corso
    })
    n_stampe += 1

# 5. Prepara intestazioni azienda (SARCA) per la lettura
headers_azienda = [
    cell.value.lower() if cell.value is not None else ""
    for cell in next(sheet_azienda.iter_rows(min_row=1, max_row=1))
]
# 6. Crea mapping per i campi corrispondenti, aggiorna secondo i nomi corretti
mapping = {
    'cognome': 'nominativo',  # qui forse devi integrare nome e cognome, oppure usa solo cognome
    'codice fiscale': 'codice_fiscale',
    'comune': 'comune',
    'provincia': 'provincia',
    'sesso': 'sesso',
    'e-mail': 'e_mail',
    # id_azienda sarà sempre azienda_da_gen
}

# 7. Trova indici master per scrivere dati
idx_cols_master = {col: headers_master.index(col) for col in headers_master}

# 8. Itera sui record azienda e inserisci solo quelli non presenti
righe_aggiunte = 0
for row in sheet_azienda.iter_rows(min_row=2, values_only=True):
    record = {headers_azienda[i]: row[i] for i in range(len(headers_azienda))}
    cf = record.get('codice fiscale')
    if cf is None or cf in data:
        continue  # salta se codice fiscale mancante o già presente

    # Prepara nuova riga per master
    new_row = [None] * len(headers_master)
    
    comune_value, provincia_value = decodifica_da_dizionario(cf, dict_codici_catastali)

    if comune_value and provincia_value == None:
        comune_value = "non trovato"
        provincia_value = "non trovato"
        print("R-{row} - CF {cf}: Codice catastale non trovato o errore nel file.")
        logging.info("R-{row} - CF {cf}: Codice catastale non trovato o errore nel file.")
        
    # Campi mappati dal file azienda a master
    # Dentro il ciclo di inserimento nuovi record:

    for key_azi, key_master in mapping.items():
        if key_master == 'nominativo':
            # Usa direttamente il campo 'cognome' che contiene già nome e cognome completo
            nominativo = record.get('cognome', '').strip()
            new_row[idx_cols_master[key_master]] = nominativo
        else:
            new_row[idx_cols_master[key_master]] = record.get(key_azi)
    
    # Scrivi comune e provincia sul master
    new_row[idx_cols_master['comune']] = comune_value
    new_row[idx_cols_master['provincia']] = provincia_value

    # Assegna azienda_da_gen e id_corso (puoi modificare id_corso a piacere)
    new_row[idx_cols_master['id_azienda']] = azienda_da_gen
    new_row[idx_cols_master['id_corso']] = None  # o un valore di default se vuoi

    # Aggiungi la nuova riga al foglio master
    sheet_master.append(new_row)
    logging.info(f"MASTER - add row - : {new_row}")
    righe_aggiunte += 1
    
print("+----------------------------------------------------------------------------------+")
print(f"Righe nuove aggiunte dal file azienda '{azienda_da_gen}': {righe_aggiunte}")
logging.info(f"Righe nuove aggiunte dal file azienda '{azienda_da_gen}': {righe_aggiunte}")
print("+----------------------------------------------------------------------------------+")
# 9. Salva il file master aggiornato
workbook.save(file_master_path)

# 10. Crea PDF se l'utente sceglie 1, se sceglie 2 o qualcosaltro esci

# GUI di selezione
cornice = """
+-------------------------------------------------------+
| Seleziona come proseguire, digita:                    |
| 1 - Genera PDF                                        |
| 2 - Confronto con dati Odoo                           |
|  Premi un tasto qualsiasi per uscire dal programma    | 
+-------------------------------------------------------+
Digita e premi ENTER: """

scelta = input(cornice)

if scelta == '1':
    logging.info("------------------------------- INIZIO GENERAZIONE PDF ------------------------------------ ")
    genera_PDF(data, azienda_da_gen)
    logging.info("------------------------------- PDF GENERATI ---------------------------------------------- ")
    logging.info("------------------------------- EXITING SESSION LOG --------------------------------------- ")
    logging.info("------------------------------------------------------------------------------------------- ")
elif scelta == '2':
    print("Funzione ancora non implementata...")
    print("Uscita dal programma.")
    logging.info("------------------------------- EXITING SESSION LOG --------------------------------------- ")
    logging.info("------------------------------------------------------------------------------------------- ")
else:
    print("Scelta non valida. Uscita dal programma.")
    logging.info("N. ERRORI CRITICI [except for codice_fiscale, info in data.items]: "+str(n_errori))
    
    logging.info("------------------------------- EXITING SESSION LOG --------------------------------------- ")
    logging.info("------------------------------------------------------------------------------------------- ")